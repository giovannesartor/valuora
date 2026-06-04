"""Portfolio Excel export — 3 sheets: Resumo, Detalhado, Simulações."""
from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EMERALD = "059669"
GRAY    = "94a3b8"
WHITE   = "FFFFFF"


def _header_row(ws, row: int, headers: list[str], bg: str = EMERALD, fg: str = WHITE):
    fill   = PatternFill("solid", fgColor=bg)
    font   = Font(bold=True, color=fg, name="Arial", size=9)
    border = Border(bottom=Side(style="medium", color="047857"))
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill, c.font, c.border, c.alignment = fill, font, border, align


def _row_fill(i: int):
    return PatternFill("solid", fgColor="F0FDF4" if i % 2 == 1 else WHITE)


def _title(ws, text: str, subtitle: str | None = None):
    ws.sheet_view.showGridLines = False
    ws["A1"].value = text
    ws["A1"].font = Font(bold=True, color=EMERALD, name="Arial", size=13)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    if subtitle:
        ws["A2"].value = subtitle
        ws["A2"].font = Font(color=GRAY, name="Arial", size=8)


def _col_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_portfolio_xlsx(analyses: list, simulations_by_id: dict | None = None) -> bytes:
    simulations_by_id = simulations_by_id or {}
    paid = [a for a in analyses if a.get("plan")]
    today = date.today().strftime("%d/%m/%Y")

    wb = Workbook()

    # ── Sheet 1: Resumo ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells("A1:H1")
    ws1.merge_cells("A2:H2")
    _title(ws1, "Valuora — Valuation Portfolio",
           f"Exported on {today} · {len(paid)} paid analysis")

    headers1 = ["Company", "Sector", "Plan", "Equity Value (USD)", "Risk Score", "Maturity", "Status", "Date"]
    _header_row(ws1, 4, headers1)

    for i, a in enumerate(paid, 5):
        rf = _row_fill(i)
        vals = [
            a.get("company_name", ""),
            a.get("sector", ""),
            a.get("plan", ""),
            float(a["equity_value"]) if a.get("equity_value") is not None else None,
            float(a["risk_score"])   if a.get("risk_score")   is not None else None,
            float(a.get("maturity_index") or 0) or None,
            a.get("status", ""),
            (a.get("created_at") or "")[:10],
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.font  = Font(name="Arial", size=9)
            c.fill  = rf
            c.alignment = Alignment(vertical="center")
            if col == 4 and val is not None:
                c.number_format = "$ #,##0.00"
            elif col in (5, 6) and val is not None:
                c.number_format = "0.0"

    # Total row
    last_data = len(paid) + 4
    tr = last_data + 2
    ws1.cell(row=tr, column=1, value=f"Total: {len(paid)} analysis").font = Font(bold=True, color=EMERALD, name="Arial", size=9)
    if paid:
        tc = ws1.cell(row=tr, column=4, value=f"=SUM(D5:D{last_data})")
        tc.number_format = "$ #,##0.00"
        tc.font = Font(bold=True, name="Arial", size=9)

    _col_widths(ws1, [28, 14, 14, 22, 15, 13, 14, 12])

    # ── Sheet 2: Detalhado ───────────────────────────────────
    ws2 = wb.create_sheet("Detail")
    ws2.merge_cells("A1:L1")
    ws2.merge_cells("A2:L2")
    _title(ws2, "Technical Detail", f"v7 parameters and fields for each analysis · {today}")

    headers2 = [
        "Company", "Sector", "Revenue (USD)", "Net Margin", "Growth",
        "Ke (WACC)", "DLOM", "Qual. Score", "Equity (USD)",
        "Implied Rating", "OLS Outliers", "Date",
    ]
    _header_row(ws2, 4, headers2)

    for i, a in enumerate(paid, 5):
        rf = _row_fill(i)
        vr   = a.get("valuation_result") or {}
        icr  = vr.get("implicit_credit_rating") or {}
        dlom = vr.get("dlom") or {}
        qual = vr.get("qualitative_assessment") or {}
        outliers = vr.get("outlier_years") or []
        vals = [
            a.get("company_name", ""),
            a.get("sector", ""),
            float(a["revenue"])    if a.get("revenue")    is not None else None,
            float(a["net_margin"]) if a.get("net_margin") is not None else None,
            float(a["growth_rate"])if a.get("growth_rate")is not None else None,
            float(vr["wacc"])      if vr.get("wacc")      is not None else None,
            float(dlom["dlom_pct"])if dlom.get("dlom_pct")is not None else None,
            float(qual["score"])   if qual.get("score")   is not None else None,
            float(a["equity_value"])if a.get("equity_value")is not None else None,
            (icr.get("rating") or "").upper() or "—",
            ", ".join(str(y) for y in outliers) if outliers else "—",
            (a.get("created_at") or "")[:10],
        ]
        pct_cols = {4, 5, 6, 7}
        brl_cols = {3, 9}
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font  = Font(name="Arial", size=9)
            c.fill  = rf
            c.alignment = Alignment(vertical="center")
            if col in brl_cols and val is not None:
                c.number_format = "$ #,##0.00"
            elif col in pct_cols and val is not None:
                c.number_format = "0.0%"

    _col_widths(ws2, [28, 14, 18, 12, 12, 10, 10, 12, 22, 14, 14, 12])

    # ── Sheet 3: Simulations ─────────────────────────────────
    ws3 = wb.create_sheet("Simulations")
    ws3.merge_cells("A1:G1")
    ws3.merge_cells("A2:G2")
    _title(ws3, "Simulation History", f"All scenario simulations for paid analyses · {today}")

    headers3 = ["Company", "Growth", "Net Margin", "Custom Ke", "Simulated Equity (USD)", "Δ vs Base (USD)", "Date"]
    _header_row(ws3, 4, headers3)

    row_idx = 5
    for a in paid:
        sims = simulations_by_id.get(str(a.get("id")), [])
        base_eq = float(a.get("equity_value") or 0)
        for s in sims:
            params  = s.get("parameters") or {}
            sim_eq  = float(s.get("equity_value") or 0)
            rf = _row_fill(row_idx)
            vals = [
                a.get("company_name", ""),
                float(params["growth_rate"])   if params.get("growth_rate")   is not None else None,
                float(params["net_margin"])    if params.get("net_margin")    is not None else None,
                float(params["discount_rate"]) if params.get("discount_rate") is not None else None,
                sim_eq   or None,
                (sim_eq - base_eq) or None,
                (s.get("created_at") or "")[:10],
            ]
            for col, val in enumerate(vals, 1):
                c = ws3.cell(row=row_idx, column=col, value=val)
                c.font  = Font(name="Arial", size=9)
                c.fill  = rf
                c.alignment = Alignment(vertical="center")
                if col in (2, 3, 4) and val is not None:
                    c.number_format = "0.0%"
                elif col in (5, 6) and val is not None:
                    c.number_format = '$ #,##0.00;[Red]($ #,##0.00);"-"'
            row_idx += 1

    if row_idx == 5:
        c = ws3.cell(row=5, column=1, value="Nenhuma simulação registrada.")
        c.font = Font(color=GRAY, name="Arial", size=9, italic=True)

    _col_widths(ws3, [28, 14, 14, 12, 24, 24, 12])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

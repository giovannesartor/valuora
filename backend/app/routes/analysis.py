import uuid
import os
import re
import asyncio
import logging
from datetime import datetime, timezone, timedelta
from pathlib import Path as FilePath
from typing import Any, Dict, List, Optional
from pydantic import BaseModel, Field
from fastapi import APIRouter, Depends, BackgroundTasks, UploadFile, File, Form, HTTPException, Query, Request, Body, Header
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.core.cache import cache_get, cache_set, cache_delete_pattern, CACHE_TTL_SHORT
from app.core.audit import audit_log
from app.core.valuation_engine.engine import run_valuation, run_valuation_with_ibge
from app.core.valuation_engine.sectors import get_sector_list

# Prevent background tasks from being garbage-collected mid-execution
_bg_tasks: set = set()

def _fire_and_forget(coro):
    """Schedule a coroutine as a background task that won't be GC'd."""
    task = asyncio.create_task(coro)
    _bg_tasks.add(task)
    task.add_done_callback(_bg_tasks.discard)
from app.services.sector_analysis_service import (
    get_dcf_sector_adjustment, _sector_to_cnae,
)
from app.services.deepseek_service import (
    extract_financial_data, generate_strategic_analysis, estimate_sector_data_with_ai,
    get_ma_comparables, detect_data_inconsistencies, analyze_historical_financials,
)
from app.models.models import (
    User, Analysis, AnalysisVersion, UserFavorite,
    AnalysisStatus, PlanType, Report, UserFeedback,
)
from app.schemas.analysis import (
    AnalysisCreate, AnalysisResponse, AnalysisListResponse,
    PaginatedAnalysesResponse,
)
from app.schemas.auth import MessageResponse

logger = logging.getLogger(__name__)
from app.services.auth_service import get_current_user
from app.services.storage_service import save_logo as _storage_save_logo
from app.services.email_service import send_report_ready_email

router = APIRouter(prefix="/analyses", tags=["Análises"])


# ─── Document validation helpers ─────────────────────────────────────────────

VALID_DOC_TYPES = [
    'DRE',
    'Balanço Patrimonial',
    'Balancete',
    'Fluxo de Caixa',
    'Contrato de Dívida',
    'Outros',
]

def _infer_from_filename(filename: str):
    """Infer document type and fiscal year from filename.
    Returns (doc_type: str | None, fiscal_year: int | None).
    doc_type is one of the VALID_DOC_TYPES values, or None if uncertain.
    """
    name = filename.lower()
    year_match = re.search(r'\b(20\d{2})\b', filename)
    year = int(year_match.group(1)) if year_match else None

    if 'balancete' in name:
        doc_type = 'Balancete'
    elif any(kw in name for kw in ('balanço patrimonial', 'balanco patrimonial',
                                    'balance sheet', 'balanço', 'balanco',
                                    'patrimoni', 'balance_', 'bp_', '_bp')):
        doc_type = 'Balanço Patrimonial'
    elif any(kw in name for kw in ('dre', 'demonstracao', 'demonstração',
                                    'resultado', 'income', 'p&l', 'profit')):
        doc_type = 'DRE'
    elif any(kw in name for kw in ('fluxo', 'caixa', 'cashflow', 'cash_flow', 'fcf')):
        doc_type = 'Fluxo de Caixa'
    elif any(kw in name for kw in ('divida', 'dívida', 'contrato', 'financiamento',
                                    'emprestimo', 'empréstimo', 'credito', 'crédito',
                                    'debt', 'loan', 'cce', 'ccb')):
        doc_type = 'Contrato de Dívida'
    else:
        doc_type = None  # uncertain — frontend will show "Não identificado" and ask user

    return doc_type, year


def _validate_document_set(file_results: list, explicit_labels: list = None) -> list:
    """Validate uploaded document set against business rules.

    Rules:
      1. All identified years must be within [current_year - 3, current_year].
      2. Max 1 DRE + 1 Balanço Patrimonial per year.

    When explicit_labels is provided (list of {"type": str, "year": int|str} dicts,
    one per file), those values are used directly — AI/filename inference is skipped.

    Args:
        file_results: [(filename, extraction_dict), ...]
        explicit_labels: Optional list of {"type": ..., "year": ...} dicts.

    Returns:
        List of human-readable error strings. Empty list = valid.
    """
    import json as _json  # local import to avoid shadowing top-level json usage

    current_year = datetime.now(timezone.utc).year
    min_year = current_year - 3

    errors: list = []
    # {year: {'DRE': [fname], 'Balanço Patrimonial': [fname]}}
    docs_by_year: Dict[int, Dict[str, List[str]]] = {}

    for idx, (filename, data) in enumerate(file_results):
        if 'error' in data:
            continue

        # ── Determine doc_type and doc_year ──────────────────────────────────
        if explicit_labels and idx < len(explicit_labels):
            label = explicit_labels[idx] or {}
            doc_type = label.get('type') or None
            raw_year = label.get('year')
            try:
                doc_year = int(raw_year) if raw_year else None
            except (TypeError, ValueError):
                doc_year = None
        else:
            fname_type, fname_year = _infer_from_filename(filename)

            # AI classification takes priority
            ai_type_raw = (data.get('document_type') or '').strip()
            ai_year_raw = data.get('fiscal_year')

            # Normalise AI type to canonical buckets
            al = ai_type_raw.lower()
            if 'balancete' in al:
                ai_type = 'Balanço Patrimonial'   # balancete substitui balanço
            elif any(kw in al for kw in ('dre', 'demonstração do resultado',
                                         'demonstracao do resultado',
                                         'income statement', 'resultado do exercício',
                                         'resultado do exercicio')):
                ai_type = 'DRE'
            elif any(kw in al for kw in ('balanço patrimonial', 'balanco patrimonial',
                                          'balance sheet', 'balanço', 'balanco')):
                ai_type = 'Balanço Patrimonial'
            else:
                ai_type = None

            # Normalise filename type (balancete → balanço)
            if fname_type == 'Balancete':
                fname_type = 'Balanço Patrimonial'

            doc_type = ai_type or fname_type
            doc_year = int(ai_year_raw) if ai_year_raw else fname_year

        if not doc_year:
            continue   # can't determine year — skip silently

        # Rule 1: year range
        if doc_year < min_year or doc_year > current_year:
            errors.append(
                f'"{filename}": exercício {doc_year} fora do intervalo permitido '
                f'({min_year}–{current_year}). Envie documentos dos últimos 3 anos.'
            )
            continue

        if doc_type in ('DRE', 'Balanço Patrimonial'):
            if doc_year not in docs_by_year:
                docs_by_year[doc_year] = {'DRE': [], 'Balanço Patrimonial': []}
            docs_by_year[doc_year][doc_type].append(filename)

    if errors:
        return errors

    # Rule 2: max 1 per type per year
    for year, types in sorted(docs_by_year.items()):
        if len(types['DRE']) > 1:
            extras = ', '.join(f'"{f}"' for f in types['DRE'][1:])
            errors.append(f'Ano {year}: máximo 1 DRE por ano. Remova: {extras}.')
        if len(types['Balanço Patrimonial']) > 1:
            extras = ', '.join(f'"{f}"' for f in types['Balanço Patrimonial'][1:])
            errors.append(f'Ano {year}: máximo 1 Balanço Patrimonial por ano. Remova: {extras}.')

    return errors


# ─────────────────────────────────────────────────────────────────────────────

ALLOWED_LOGO_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/svg+xml", "image/webp"}
MAX_LOGO_SIZE = settings.MAX_LOGO_SIZE_BYTES


async def _save_logo(logo: UploadFile, analysis_id: uuid.UUID) -> str:
    """Valida e persiste o logo — usa R2 se configurado, filesystem local como fallback."""
    if logo.content_type not in ALLOWED_LOGO_TYPES:
        raise HTTPException(status_code=400, detail="Formato de logo não suportado. Use PNG, JPG, SVG ou WebP.")
    content = await logo.read()
    if len(content) > MAX_LOGO_SIZE:
        raise HTTPException(status_code=400, detail="Logo deve ter no máximo 2MB.")
    ext = logo.filename.rsplit(".", 1)[-1].lower() if logo.filename else "png"
    return await _storage_save_logo(content, analysis_id, ext)


# ─── Per-user rate limiter for expensive operations ───────────────────────────
_USER_ANALYSIS_LIMIT = 10    # max analysis creations
_USER_ANALYSIS_WINDOW = 3600  # per 1 hour (seconds)


async def _check_user_analysis_rate_limit(user_id: str) -> None:
    """Raise 429 if the user has exceeded their hourly analysis creation quota.
    Falls back to allowing requests if Redis is unavailable."""
    from app.core.redis import redis_client
    key = f"rl:user_analysis:{user_id}"
    try:
        current = await redis_client.incr(key)
        if current == 1:
            await redis_client.expire(key, _USER_ANALYSIS_WINDOW)
        if current > _USER_ANALYSIS_LIMIT:
            raise HTTPException(
                status_code=429,
                detail=f"Limite de {_USER_ANALYSIS_LIMIT} análises por hora atingido. Tente novamente mais tarde.",
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"[RateLimit] Redis unavailable for user rate limit: {e!r} — skipping")


@router.post("/", response_model=AnalysisResponse)
async def create_analysis(
    data: AnalysisCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check_user_analysis_rate_limit(str(current_user.id))
    if data.revenue <= 0:
        raise HTTPException(status_code=400, detail="Receita deve ser maior que zero.")
    analysis = Analysis(
        user_id=current_user.id,
        partner_id=current_user.partner_id,  # propagate referral tracking
        company_name=data.company_name,
        sector=data.sector,
        cnpj=data.cnpj,
        revenue=data.revenue,
        net_margin=data.net_margin,
        growth_rate=data.growth_rate,
        debt=data.debt,
        cash=data.cash,
        founder_dependency=data.founder_dependency,
        projection_years=data.projection_years,
        ebitda=data.ebitda,
        recurring_revenue_pct=data.recurring_revenue_pct,
        num_employees=data.num_employees,
        years_in_business=data.years_in_business,
        previous_investment=data.previous_investment,
        qualitative_answers=data.qualitative_answers,
        dcf_weight=data.dcf_weight,
        custom_exit_multiple=data.custom_exit_multiple,
        company_type=data.company_type,
        revenue_ntm=data.revenue_ntm,
        ebitda_margin=data.ebitda_margin,
        tangible_assets=data.tangible_assets,
        intangible_assets=data.intangible_assets,
        equity_participations=data.equity_participations,
        monthly_burn_rate=data.monthly_burn_rate,
        pending_assets=[a.model_dump() for a in data.pending_assets] if data.pending_assets else None,
        status=AnalysisStatus.PROCESSING,
    )
    db.add(analysis)
    await db.flush()

    # Fetch IBGE sector adjustment (non-blocking, with fallback chain)
    # 1. IBGE/SIDRA (melhor) → 2. DeepSeek AI (bom) → 3. Damodaran estático (seguro)
    ibge_adj = None
    try:
        cnae_code = _sector_to_cnae(data.sector)
        adjustment = await get_dcf_sector_adjustment(
            cnae_code=cnae_code,
            company_revenue=float(data.revenue),
            company_growth=data.growth_rate,
        )
        ibge_adj = adjustment.model_dump()
    except Exception as e:
        logger.warning(f"[ANALYSIS] IBGE adjustment failed for {data.sector}: {e}")
        # IBGE falhou → tentar DeepSeek como fallback
        try:
            cnae_code = _sector_to_cnae(data.sector)
            ai_sector = await estimate_sector_data_with_ai(data.sector, cnae_code)
            if ai_sector:
                ibge_adj = ai_sector
        except Exception as e2:
            logger.warning(f"[ANALYSIS] DeepSeek sector fallback failed: {e2}")

    # Run valuation with IBGE data if available (CPU-bound → offload to thread)
    _engine_kwargs = dict(
        years_in_business=data.years_in_business,
        ebitda=float(data.ebitda) if data.ebitda else None,
        recurring_revenue_pct=data.recurring_revenue_pct,
        num_employees=data.num_employees,
        previous_investment=float(data.previous_investment),
        qualitative_answers=data.qualitative_answers,
        dcf_weight=data.dcf_weight,
        custom_exit_multiple=data.custom_exit_multiple,
        # v5 diagnostic fields
        company_type=data.company_type,
        revenue_ntm=float(data.revenue_ntm) if data.revenue_ntm else None,
        ebitda_margin=data.ebitda_margin,
        tangible_assets=float(data.tangible_assets) if data.tangible_assets else 0,
        intangible_assets=float(data.intangible_assets) if data.intangible_assets else 0,
        equity_participations=float(data.equity_participations) if data.equity_participations else 0,
        monthly_burn_rate=float(data.monthly_burn_rate) if data.monthly_burn_rate else None,
        pending_assets=[a.model_dump() for a in data.pending_assets] if data.pending_assets else None,
    )
    try:
        if ibge_adj:
            result = await asyncio.to_thread(
                run_valuation_with_ibge,
                revenue=float(data.revenue),
                net_margin=data.net_margin,
                sector=data.sector,
                ibge_adjustment=ibge_adj,
                growth_rate=data.growth_rate,
                debt=float(data.debt),
                cash=float(data.cash),
                founder_dependency=data.founder_dependency,
                projection_years=data.projection_years,
                **_engine_kwargs,
            )
        else:
            result = await asyncio.to_thread(
                run_valuation,
                revenue=float(data.revenue),
                net_margin=data.net_margin,
                sector=data.sector,
                growth_rate=data.growth_rate,
                debt=float(data.debt),
                cash=float(data.cash),
                founder_dependency=data.founder_dependency,
                projection_years=data.projection_years,
                **_engine_kwargs,
            )
    except Exception as engine_err:
        from app.core.observability import report_exc
        report_exc(engine_err, "analysis.valuation_engine", flow="manual", analysis_id=str(getattr(analysis, 'id', None)))
        analysis.status = AnalysisStatus.FAILED
        await db.commit()
        raise HTTPException(status_code=500, detail="Erro no motor de valuation. Tente novamente.")

    analysis.valuation_result = result
    analysis.equity_value = result["equity_value"]
    analysis.risk_score = result["risk_score"]
    analysis.maturity_index = result["maturity_index"]
    analysis.percentile = result["percentile"]
    analysis.status = AnalysisStatus.COMPLETED

    # Save version
    version = AnalysisVersion(
        analysis_id=analysis.id,
        version_number=1,
        valuation_result=result,
        equity_value=result["equity_value"],
    )
    db.add(version)

    await db.commit()
    await db.refresh(analysis)
    # Invalidate KPI and sidebar caches for this user
    await cache_delete_pattern(f"qv:kpis:{current_user.id}")
    await cache_delete_pattern(f"qv:sidebar:{current_user.id}")
    # Notify user by email
    _fire_and_forget(send_report_ready_email(
        current_user.email,
        current_user.full_name or current_user.email,
        analysis.company_name,
        f"{settings.FRONTEND_URL}/analise/{analysis.id}",
    ))
    return analysis


@router.post("/extract-preview")
async def extract_preview(
    files: List[UploadFile] = File(...),
    file_labels: Optional[str] = Form(None),  # JSON: [{"type": "DRE", "year": 2025}, ...]
    current_user: User = Depends(get_current_user),
):
    """Extrai dados financeiros dos arquivos enviados sem criar a análise.
    Usado para mostrar o painel de preview antes das perguntas qualitativas."""
    if not files:
        raise HTTPException(status_code=400, detail="Envie pelo menos um arquivo.")
    if len(files) > 15:
        raise HTTPException(status_code=400, detail="Máximo de 15 arquivos permitidos.")

    # Parse explicit file labels if provided
    import json as _json_prev
    _explicit_labels = None
    if file_labels:
        try:
            _explicit_labels = _json_prev.loads(file_labels)
        except Exception:
            _explicit_labels = None

    _MAX_FILE_SIZE = 15 * 1024 * 1024  # 15 MB per file
    file_contents = []
    for file in files:
        ext = file.filename.split(".")[-1].lower() if file.filename else ""
        if ext not in ("pdf", "xlsx", "xls"):
            raise HTTPException(status_code=400, detail=f"Formato não suportado: {file.filename}. Envie PDF ou Excel.")
        content = await file.read()
        if len(content) > _MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail=f"Arquivo {file.filename!r} excede o limite de 15 MB.")
        file_contents.append((file.filename, content, ext))

    async def _extract_one(filename: str, content: bytes, ext: str):
        try:
            return await extract_financial_data(content, ext)
        except Exception as e:
            logger.warning(f"[EXTRACT-PREVIEW] Extraction failed for {filename}: {e}")
            return {"error": str(e)}

    extraction_results = await asyncio.gather(
        *[_extract_one(name, content, ext) for name, content, ext in file_contents]
    )

    # ── Validate document set (year range, duplicates) ──
    _val_input = [(fname, res) for (fname, _, _), res in zip(file_contents, extraction_results)]
    _val_errors = _validate_document_set(_val_input, _explicit_labels)
    if _val_errors:
        raise HTTPException(
            status_code=422,
            detail=_val_errors if len(_val_errors) > 1 else _val_errors[0],
        )

    # Sort by year descending: prefer explicit label year, fall back to AI fiscal_year
    def _sort_year(item):
        idx, ((_, _, _), result_data) = item
        if _explicit_labels and idx < len(_explicit_labels):
            lbl = _explicit_labels[idx] or {}
            try:
                return int(lbl.get('year') or 0)
            except (TypeError, ValueError):
                pass
        return int(result_data.get("fiscal_year") or 0) if "error" not in result_data else 0

    _paired = list(enumerate(zip(file_contents, extraction_results)))
    _paired.sort(key=_sort_year, reverse=True)
    _paired = [item for _, item in _paired]

    merged: dict = {}
    sources: dict = {}  # campo -> nome do arquivo de origem
    all_notes: list = []
    for (filename, _, _), result_data in _paired:
        if "error" not in result_data:
            if result_data.get("notes"):
                all_notes.append(result_data["notes"])
            for k, v in result_data.items():
                if k == "notes":
                    continue
                # Most-recent year is processed first; only fill truly missing/empty slots
                if v is not None and v != "" and (k not in merged or merged[k] is None or merged[k] == ""):
                    merged[k] = v
                    sources[k] = filename
    if all_notes:
        merged["notes"] = " | ".join(all_notes)

    if not merged:
        raise HTTPException(
            status_code=422,
            detail="Não foi possível extrair dados dos documentos enviados. Verifique se os arquivos contêm DRE ou Balanço Patrimonial legíveis.",
        )

    # Build per-file identification summary for frontend
    # Each file gets: filename, ai_identified_type (from extraction), inferred_type (from filename), year
    _file_identifications = []
    for i, (filename, _, _) in enumerate(file_contents):
        res = extraction_results[i]
        ai_type = res.get("document_type") if "error" not in res else None
        fname_type, fname_year = _infer_from_filename(filename)
        ai_year = res.get("fiscal_year") if "error" not in res else None

        # Determine best type: explicit label > AI > filename inference
        explicit_type = None
        explicit_year = None
        if _explicit_labels and i < len(_explicit_labels):
            lbl = _explicit_labels[i] or {}
            explicit_type = lbl.get("type")
            explicit_year = lbl.get("year")

        best_type = explicit_type or ai_type or fname_type  # None = "Não identificado"
        best_year = explicit_year or ai_year or fname_year
        confident = bool(explicit_type or ai_type)  # False = show "Não identificado" UI

        _file_identifications.append({
            "filename": filename,
            "identified_type": best_type,
            "identified_year": best_year,
            "confident": confident,  # if False, frontend shows manual selector
            "ai_type": ai_type,
            "fname_type": fname_type,
        })

    # Conta campos reais extraídos (ignora metadados internos)
    _data_keys = [k for k in merged if not k.startswith("_") and k != "notes" and k != "error" and merged[k] is not None]
    return {
        **merged,
        "_sources": sources,
        "_file_count": len(file_contents),
        "_field_count": len(_data_keys),
        "_file_identifications": _file_identifications,
        "_valid_doc_types": VALID_DOC_TYPES,
    }


@router.post("/validate-data")
async def validate_financial_data(
    data: Dict[str, Any] = Body(...),
    sector: str = Query(...),
    current_user: User = Depends(get_current_user),
):
    """Detecta inconsistências e outliers nos dados financeiros antes de confirmar a geração.

    Retorna lista de warnings com severidade para exibição no frontend antes de criar a análise.
    """
    result = await detect_data_inconsistencies(data, sector)
    return result


@router.post("/upload", response_model=AnalysisResponse)
async def create_analysis_from_upload(
    company_name: str = Form(...),
    sector: str = Form(...),
    cnpj: str = Form(""),
    founder_dependency: float = Form(0.0),
    projection_years: int = Form(10),
    qualitative_answers: Optional[str] = Form(None),
    file_labels: Optional[str] = Form(None),  # JSON: [{"type": "DRE", "year": 2025}, ...]
    pending_assets: Optional[str] = Form(None),  # JSON: [{label, asset_type, value, confidence, mode, ...}]
    analysis_objective: Optional[str] = Form(None),  # "captacao" | "venda" | "socio" | None
    files: List[UploadFile] = File(...),
    logo: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Cria análise a partir de upload de DRE/Balanço (PDF ou Excel). Aceita múltiplos arquivos."""
    await _check_user_analysis_rate_limit(str(current_user.id))
    logger.info(
        "[UPLOAD] Start: company=%r, sector=%r, files=%d, logo=%s",
        company_name, sector,
        len(files) if files else 0,
        'yes' if logo and logo.filename else 'no',
    )
    if not files:
        raise HTTPException(status_code=400, detail="Envie pelo menos um arquivo.")
    if len(files) > 15:
        raise HTTPException(status_code=400, detail="Máximo de 15 arquivos permitidos.")

    # Parse founder_dependency from percentage to decimal
    founder_dep = min(max(founder_dependency / 100, 0), 1)

    # Parse qualitative_answers JSON
    qual_answers = None
    if qualitative_answers:
        try:
            import json as _json
            qual_answers = _json.loads(qualitative_answers)
        except Exception:
            qual_answers = None

    # Parse explicit file labels
    import json as _json_up
    _explicit_labels_up = None
    if file_labels:
        try:
            _explicit_labels_up = _json_up.loads(file_labels)
        except Exception:
            _explicit_labels_up = None

    # ── Read all file contents and validate extensions ──
    MAX_FILE_SIZE = 15 * 1024 * 1024  # 15 MB per file
    file_contents = []
    for file in files:
        ext = file.filename.split(".")[-1].lower() if file.filename else ""
        if ext not in ("pdf", "xlsx", "xls"):
            raise HTTPException(status_code=400, detail=f"Formato não suportado: {file.filename}. Envie PDF ou Excel.")
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=413, detail=f"Arquivo {file.filename!r} excede o limite de 15 MB.")
        file_contents.append((file.filename, content, ext))

    # ── Extract financial data from ALL files concurrently ──
    # Each DeepSeek call can take up to 60s; sequential = N*60s, concurrent ≈ 60s
    async def _extract_one(filename: str, content: bytes, ext: str):
        try:
            return await extract_financial_data(content, ext)
        except Exception as e:
            logger.warning(f"[UPLOAD] Extraction failed for {filename}: {e}")
            return {"error": str(e)}

    logger.info("[UPLOAD] Extracting data from %d files concurrently...", len(file_contents))
    extraction_results = await asyncio.gather(
        *[_extract_one(name, content, ext) for name, content, ext in file_contents]
    )

    # ── Validate document set (year range, duplicates) ──
    _val_input_up = [(fname, res) for (fname, _, _), res in zip(file_contents, extraction_results)]
    _val_errors_up = _validate_document_set(_val_input_up, _explicit_labels_up)
    if _val_errors_up:
        raise HTTPException(
            status_code=422,
            detail=_val_errors_up if len(_val_errors_up) > 1 else _val_errors_up[0],
        )

    # Sort by year descending: prefer explicit label year, fall back to AI fiscal_year
    def _sort_year_up(item):
        idx, ((_, _, _), result_data) = item
        if _explicit_labels_up and idx < len(_explicit_labels_up):
            lbl = _explicit_labels_up[idx] or {}
            try:
                return int(lbl.get('year') or 0)
            except (TypeError, ValueError):
                pass
        return int(result_data.get("fiscal_year") or 0) if "error" not in result_data else 0

    _paired_upload = list(enumerate(zip(file_contents, extraction_results)))
    _paired_upload.sort(key=_sort_year_up, reverse=True)
    _paired_upload = [item for _, item in _paired_upload]

    all_extracted = {}
    all_upload_notes: list = []
    uploaded_filenames = []
    # Item 7: per-year historical data collection
    _year_data: dict = {}
    for (filename, _, _), result_data in _paired_upload:
        if "error" not in result_data:
            if result_data.get("notes"):
                all_upload_notes.append(result_data["notes"])
            for k, v in result_data.items():
                if k == "notes":
                    continue
                # Most-recent year processed first; only fill missing/empty slots from older years
                if v is not None and v != "" and (k not in all_extracted or all_extracted[k] is None or all_extracted[k] == ""):
                    all_extracted[k] = v
            # Collect per-year financials for historical series
            _fy = result_data.get("fiscal_year") or result_data.get("year")
            if _fy:
                try:
                    _fy = int(_fy)
                except (TypeError, ValueError):
                    _fy = None
            if _fy:
                _year_data[_fy] = {
                    "revenue": result_data.get("revenue"),
                    "net_margin": result_data.get("net_margin"),
                    "ebitda": result_data.get("ebitda"),
                    "debt": result_data.get("total_liabilities"),
                    "cash": result_data.get("cash"),
                }
        uploaded_filenames.append(filename)
    if all_upload_notes:
        all_extracted["notes"] = " | ".join(all_upload_notes)

    # Item 7: build historical arrays if >= 2 years found
    _historical_revenues_upload: list = []
    _historical_margins_upload: list = []
    _ai_historical_analysis = None
    if len(_year_data) >= 2:
        for _yr in sorted(_year_data.keys()):
            _yd = _year_data[_yr]
            if _yd.get("revenue") and float(_yd["revenue"]) > 0:
                _historical_revenues_upload.append(float(_yd["revenue"]))
                _historical_margins_upload.append(float(_yd.get("net_margin") or 0.10))

        # ── IA analisa série histórica completa: CAGR, tendências, projeções ──
        try:
            _ai_historical_analysis = await analyze_historical_financials(_year_data)
            if _ai_historical_analysis:
                all_extracted["ai_historical_analysis"] = _ai_historical_analysis
                # Se a IA sugerir uma taxa de crescimento melhor (baseada em dados reais), usar
                _ai_growth = _ai_historical_analysis.get("recommended_growth_rate")
                if _ai_growth is not None and isinstance(_ai_growth, (int, float)):
                    logger.info(
                        "[UPLOAD] AI historical analysis: CAGR=%s, recommended_growth=%s",
                        _ai_historical_analysis.get("cagr_revenue"),
                        _ai_growth,
                    )
        except Exception as _hist_err:
            logger.warning("[UPLOAD] AI historical analysis failed (non-fatal): %s", _hist_err)

    # Item 5: DRE × Balance cross-validation
    _data_quality_warnings: list = []
    if len(_year_data) >= 1:
        # Check if most-recent year has both DRE and balance fields
        _latest_yr = max(_year_data.keys())
        _ld = _year_data[_latest_yr]
        if _ld.get("revenue") and _ld.get("net_margin") and _ld.get("debt") is not None:
            _net_income_est = float(_ld["revenue"]) * float(_ld["net_margin"])
            # If we have two years, check equity delta ≈ net income
            if len(_year_data) >= 2:
                _prev_yr = sorted(_year_data.keys())[-2]
                _pd = _year_data[_prev_yr]
                _eq_curr = float(_ld.get("cash") or 0) - float(_ld.get("debt") or 0)
                _eq_prev = float(_pd.get("cash") or 0) - float(_pd.get("debt") or 0)
                _delta_eq = _eq_curr - _eq_prev
                _baseline = max(abs(_net_income_est), abs(_delta_eq), 1.0)
                if abs(_net_income_est - _delta_eq) / _baseline > 0.35:
                    _data_quality_warnings.append(
                        "Inconsistência DRE × Balanço: variação de PL diverge do lucro líquido estimado (>35%). "
                        "Verifique se os documentos pertencem ao mesmo período."
                    )
    if _data_quality_warnings:
        all_extracted["data_quality_warnings"] = _data_quality_warnings

    # Item 6: Non-recurring EBITDA normalization
    _raw_ebitda = all_extracted.get("ebitda")
    _raw_revenue = all_extracted.get("revenue") or 0
    if _raw_ebitda and _raw_revenue > 0:
        _ebitda_margin = float(_raw_ebitda) / float(_raw_revenue)
        _sector_avg_margin = 0.12  # conservative default
        if _ebitda_margin > _sector_avg_margin * 1.8:
            _normalized_ebitda = float(_raw_ebitda) * 0.85
            all_extracted["ebitda_raw"] = float(_raw_ebitda)
            all_extracted["ebitda"] = round(_normalized_ebitda, 2)
            all_extracted["ebitda_normalization_note"] = (
                f"EBITDA normalizado: margem {_ebitda_margin:.1%} > 1.8× média setorial. "
                "Possível item não-recorrente removido (–15%). Revise com o cliente."
            )
            if _data_quality_warnings is not None:
                _data_quality_warnings.append(all_extracted["ebitda_normalization_note"])

    logger.info("[UPLOAD] Extraction done. Keys found: %s", list(all_extracted.keys()))

    extracted = all_extracted
    if not extracted:
        raise HTTPException(status_code=422, detail="Não foi possível extrair dados dos documentos enviados.")

    revenue = extracted.get("revenue") or 0
    net_margin = extracted.get("net_margin") or 0.10
    growth_rate = extracted.get("growth_rate") or 0.10
    debt = extracted.get("total_liabilities") or 0
    cash = extracted.get("cash") or 0

    # Fix #9: Validar números extraídos pela IA
    if revenue <= 0:
        raise HTTPException(status_code=422, detail="Não foi possível extrair receita válida do documento.")
    if not (0 <= net_margin < 1):
        # Se veio como porcentagem (ex: 15 ao invés de 0.15)
        if 1 <= net_margin <= 100:
            net_margin = net_margin / 100
        elif net_margin < 0:
            net_margin = 0.0  # Empresa com margem negativa → breakeven
        else:
            net_margin = 0.10  # Fallback
    if not (-0.5 < growth_rate < 5):
        if 1 <= growth_rate <= 100:
            growth_rate = growth_rate / 100
        else:
            growth_rate = 0.10
    if debt < 0:
        debt = 0
    if cash < 0:
        cash = 0

    _pending_assets_list = None
    if pending_assets:
        try:
            import json as _json_pa
            _pending_assets_list = _json_pa.loads(pending_assets)
        except Exception:
            _pending_assets_list = None

    analysis = Analysis(
        user_id=current_user.id,
        partner_id=current_user.partner_id,  # propagate referral tracking (upload route)
        company_name=company_name,
        sector=sector,
        cnpj=cnpj or None,
        revenue=revenue,
        net_margin=net_margin,
        growth_rate=growth_rate,
        debt=debt,
        cash=cash,
        founder_dependency=founder_dep,
        projection_years=projection_years,
        qualitative_answers=qual_answers,
        pending_assets=_pending_assets_list or None,
        extracted_data=extracted,
        uploaded_files=uploaded_filenames,
        status=AnalysisStatus.PROCESSING,
    )
    db.add(analysis)
    await db.flush()

    # Fallback chain: IBGE → DeepSeek AI → Damodaran estático
    ibge_adj = None
    try:
        cnae_code = _sector_to_cnae(sector)
        adjustment = await get_dcf_sector_adjustment(
            cnae_code=cnae_code,
            company_revenue=float(revenue),
            company_growth=float(growth_rate),
        )
        ibge_adj = adjustment.model_dump()
    except Exception as e:
        logger.warning(f"[UPLOAD] IBGE adjustment failed for {sector}: {e}")
        # IBGE falhou → tentar DeepSeek como fallback
        try:
            cnae_code = _sector_to_cnae(sector)
            ai_sector = await estimate_sector_data_with_ai(sector, cnae_code)
            if ai_sector:
                ibge_adj = ai_sector
        except Exception as e2:
            logger.warning(f"[UPLOAD] DeepSeek sector fallback failed: {e2}")

    # Enrich engine kwargs with extracted financial data
    _ebitda = extracted.get("ebitda") or None
    _num_employees = extracted.get("num_employees") or 0
    _years_in_business = extracted.get("years_in_business") or 3
    _recurring_revenue_pct = extracted.get("recurring_revenue_pct") or 0.0
    _previous_investment = extracted.get("previous_investment") or 0.0

    _engine_kwargs = dict(
        qualitative_answers=qual_answers,
        ebitda=float(_ebitda) if _ebitda else None,
        num_employees=int(_num_employees),
        years_in_business=int(_years_in_business),
        recurring_revenue_pct=float(_recurring_revenue_pct),
        previous_investment=float(_previous_investment),
        # v5 diagnostic fields from the analysis record
        company_type=analysis.company_type,
        revenue_ntm=float(analysis.revenue_ntm) if analysis.revenue_ntm else None,
        ebitda_margin=analysis.ebitda_margin,
        tangible_assets=float(analysis.tangible_assets) if analysis.tangible_assets else 0,
        intangible_assets=float(analysis.intangible_assets) if analysis.intangible_assets else 0,
        equity_participations=float(analysis.equity_participations) if analysis.equity_participations else 0,
        monthly_burn_rate=float(analysis.monthly_burn_rate) if analysis.monthly_burn_rate else None,
        pending_assets=analysis.pending_assets or None,
        # Item 7: multi-year historical series derived from uploaded documents
        historical_revenues=_historical_revenues_upload if len(_historical_revenues_upload) >= 2 else None,
        historical_margins=_historical_margins_upload if len(_historical_margins_upload) >= 2 else None,
    )

    try:
        if ibge_adj:
            result = await asyncio.to_thread(
                run_valuation_with_ibge,
                revenue=float(revenue),
                net_margin=float(net_margin),
                sector=sector,
                ibge_adjustment=ibge_adj,
                growth_rate=float(growth_rate),
                debt=float(debt),
                cash=float(cash),
                founder_dependency=founder_dep,
                projection_years=projection_years,
                **_engine_kwargs,
            )
        else:
            result = await asyncio.to_thread(
                run_valuation,
                revenue=float(revenue),
                net_margin=float(net_margin),
                sector=sector,
                growth_rate=float(growth_rate),
                debt=float(debt),
                cash=float(cash),
                founder_dependency=founder_dep,
                projection_years=projection_years,
                **_engine_kwargs,
            )
    except Exception as engine_err:
        from app.core.observability import report_exc
        report_exc(engine_err, "analysis.valuation_engine", flow="upload", analysis_id=str(getattr(analysis, 'id', None)))
        analysis.status = AnalysisStatus.FAILED
        await db.commit()
        raise HTTPException(status_code=500, detail="Erro no motor de valuation. Tente novamente.")

    # Fix #12: AI recebe resultado do valuation para análise mais rica
    logger.info("[UPLOAD] Valuation done. Generating strategic analysis...")
    ai_text = await generate_strategic_analysis(
        extracted,
        valuation_result=result,
        sector_benchmarks=ibge_adj,
        analysis_objective=analysis_objective,
    )
    logger.info("[UPLOAD] Strategic analysis done. Saving...")

    analysis.valuation_result = result
    analysis.equity_value = result["equity_value"]
    analysis.risk_score = result["risk_score"]
    analysis.maturity_index = result["maturity_index"]
    analysis.percentile = result["percentile"]
    analysis.ai_analysis = ai_text
    analysis.status = AnalysisStatus.COMPLETED

    # Save logo if provided
    if logo and logo.filename:
        try:
            logo_rel = await _save_logo(logo, analysis.id)
            analysis.logo_path = logo_rel
        except Exception as e:
            logger.warning(f"[UPLOAD] Logo save failed for {analysis.id}: {e}")

    version = AnalysisVersion(
        analysis_id=analysis.id,
        version_number=1,
        valuation_result=result,
        equity_value=result["equity_value"],
    )
    db.add(version)

    await db.commit()
    await db.refresh(analysis)
    # Invalidate KPI and sidebar caches for this user
    await cache_delete_pattern(f"qv:kpis:{current_user.id}")
    await cache_delete_pattern(f"qv:sidebar:{current_user.id}")
    # Email is sent by _generate_and_send_report in payments.py after PDF is generated.
    # Firing here would be premature — no PDF exists yet at this point, and for regular
    # users the payment (which queues PDF generation) hasn't happened yet.
    return analysis


@router.post("/{analysis_id}/logo", response_model=MessageResponse)
async def upload_logo(
    analysis_id: uuid.UUID,
    logo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload/atualiza logo de uma análise existente."""
    analysis = (await db.execute(
        select(Analysis).where(Analysis.id == analysis_id, Analysis.user_id == current_user.id)
    )).scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    logo_rel = await _save_logo(logo, analysis.id)
    analysis.logo_path = logo_rel
    await db.commit()
    return {"message": "Logo atualizada com sucesso."}


@router.delete("/{analysis_id}", response_model=MessageResponse)
async def delete_analysis(
    analysis_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Move análise para a lixeira (soft delete — 30 dias)."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    analysis.deleted_at = datetime.now(timezone.utc)
    await db.commit()
    # Invalidate KPI and sidebar caches for this user
    await cache_delete_pattern(f"qv:kpis:{current_user.id}")
    await cache_delete_pattern(f"qv:sidebar:{current_user.id}")
    await audit_log(
        action="analysis.delete",
        user_id=str(current_user.id),
        user_email=current_user.email,
        resource_id=str(analysis_id),
        detail=f"Soft-deleted analysis: {analysis.company_name}",
        ip=request.client.host if request.client else None,
    )
    return {"message": "Análise movida para a lixeira. Será excluída permanentemente em 30 dias."}


@router.get("/trash", response_model=PaginatedAnalysesResponse)
async def list_trash(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista análises na lixeira."""
    base = select(Analysis).where(
        Analysis.user_id == current_user.id,
        Analysis.deleted_at.isnot(None),
    )
    count_q = select(func.count()).select_from(base.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    base = base.order_by(Analysis.deleted_at.desc())
    offset = (page - 1) * page_size
    items = (await db.execute(base.offset(offset).limit(page_size))).scalars().all()
    total_pages = max(1, -(-total // page_size))

    return PaginatedAnalysesResponse(
        items=items, total=total, page=page,
        page_size=page_size, total_pages=total_pages,
    )


@router.post("/{analysis_id}/restore", response_model=MessageResponse)
async def restore_analysis(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Restaura análise da lixeira."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.isnot(None),
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada na lixeira.")

    analysis.deleted_at = None
    await db.commit()
    return {"message": "Análise restaurada com sucesso."}


@router.delete("/{analysis_id}/permanent", response_model=MessageResponse)
async def permanent_delete_analysis(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exclui permanentemente uma análise da lixeira."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_not(None),  # only items already in trash can be permanently deleted
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada na lixeira.")

    # Delete logo file if exists
    if analysis.logo_path:
        logo_full = os.path.join(settings.UPLOADS_DIR, analysis.logo_path.lstrip("/"))
        if os.path.exists(logo_full):
            try:
                os.remove(logo_full)
            except OSError:
                pass

    # Delete PDF report files
    for report in (await db.execute(
        select(Report).where(Report.analysis_id == analysis_id)
    )).scalars().all():
        if report.file_path and os.path.exists(report.file_path):
            try:
                os.remove(report.file_path)
            except OSError:
                pass

    await db.delete(analysis)  # cascade deletes versions, simulations, reports, payment
    await db.commit()
    return {"message": "Análise excluída permanentemente."}


@router.get("/", response_model=PaginatedAnalysesResponse)
async def list_analyses(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    sort: str = Query("date_desc"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    base = select(Analysis).where(Analysis.user_id == current_user.id, Analysis.deleted_at.is_(None))
    if search:
        safe_search = search.replace('%', '\\%').replace('_', '\\_')  # escape LIKE wildcards
        base = base.where(Analysis.company_name.ilike(f"%{safe_search}%"))
    if status and status != "all":
        base = base.where(Analysis.status == status)
    if sector and sector != "all":
        base = base.where(Analysis.sector == sector)

    # Count
    count_q = select(func.count()).select_from(base.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    # Sort
    order_map = {
        "date_desc": Analysis.created_at.desc(),
        "date_asc": Analysis.created_at.asc(),
        "value_desc": Analysis.equity_value.desc().nullslast(),
        "value_asc": Analysis.equity_value.asc().nullsfirst(),
        "name_asc": Analysis.company_name.asc(),
        "name_desc": Analysis.company_name.desc(),
    }
    base = base.order_by(order_map.get(sort, Analysis.created_at.desc()))

    # Paginate
    offset = (page - 1) * page_size
    items = (await db.execute(base.offset(offset).limit(page_size))).scalars().all()
    total_pages = max(1, -(-total // page_size))

    return PaginatedAnalysesResponse(
        items=items, total=total, page=page,
        page_size=page_size, total_pages=total_pages,
    )


# ─── KPIs Endpoint ────────────────────────────────────────

@router.get("/kpis/summary")
async def get_kpis(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna KPIs agregados das análises do usuário. Cached 5 min per user."""
    cache_key = f"qv:kpis:{current_user.id}"
    cached = await cache_get(cache_key)
    if cached:
        return cached

    base = select(Analysis).where(
        Analysis.user_id == current_user.id,
        Analysis.deleted_at.is_(None),
    )

    # Total analyses
    total = (await db.execute(
        select(func.count()).select_from(base.subquery())
    )).scalar() or 0

    # Completed analyses
    completed_q = base.where(Analysis.status == AnalysisStatus.COMPLETED)
    completed = (await db.execute(
        select(func.count()).select_from(completed_q.subquery())
    )).scalar() or 0

    # Avg equity value
    avg_value = (await db.execute(
        select(func.avg(Analysis.equity_value)).where(
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
            Analysis.status == AnalysisStatus.COMPLETED,
            Analysis.equity_value.isnot(None),
        )
    )).scalar() or 0

    # Max equity value
    max_value = (await db.execute(
        select(func.max(Analysis.equity_value)).where(
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
            Analysis.status == AnalysisStatus.COMPLETED,
        )
    )).scalar() or 0

    # Avg risk score
    avg_risk = (await db.execute(
        select(func.avg(Analysis.risk_score)).where(
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
            Analysis.status == AnalysisStatus.COMPLETED,
            Analysis.risk_score.isnot(None),
        )
    )).scalar() or 0

    # Sector distribution
    sector_rows = (await db.execute(
        select(Analysis.sector, func.count(Analysis.id))
        .where(
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
        .group_by(Analysis.sector)
    )).all()
    sectors = {row[0]: row[1] for row in sector_rows}

    # Sparklines: last 30 days avg equity_value per day (PostgreSQL date_trunc)
    # Gracefully degrades to empty sparklines on SQLite (tests) or other DBs
    sparklines: dict = {"dates": [], "avg_value": [], "count": []}
    try:
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        sparkline_rows = (await db.execute(
            select(
                func.date_trunc("day", Analysis.created_at).label("day"),
                func.avg(Analysis.equity_value).label("avg_val"),
                func.count(Analysis.id).label("cnt"),
            )
            .where(
                Analysis.user_id == current_user.id,
                Analysis.deleted_at.is_(None),
                Analysis.status == AnalysisStatus.COMPLETED,
                Analysis.equity_value.isnot(None),
                Analysis.created_at >= thirty_days_ago,
            )
            .group_by(func.date_trunc("day", Analysis.created_at))
            .order_by(func.date_trunc("day", Analysis.created_at))
        )).all()
        sparklines = {
            "dates": [str(row.day)[:10] for row in sparkline_rows],
            "avg_value": [float(row.avg_val) for row in sparkline_rows],
            "count": [int(row.cnt) for row in sparkline_rows],
        }
    except Exception as _spark_exc:
        logger.debug("[KPIs] Sparkline query skipped: %s", _spark_exc)

    result_data = {
        "total": total,
        "completed": completed,
        "avg_value": float(avg_value),
        "max_value": float(max_value),
        "avg_risk": float(avg_risk),
        "sectors": sectors,
        "sparklines": sparklines,
    }
    await cache_set(cache_key, result_data, ttl=300)  # 5 minutes
    return result_data


# ─── CSV Export ────────────────────────────────────────────

@router.get("/export/csv")
async def export_analyses_csv(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta todas as análises do usuário em CSV."""
    import csv
    import io
    from fastapi.responses import StreamingResponse

    result = await db.execute(
        select(Analysis).where(
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        ).order_by(Analysis.created_at.desc())
    )
    analyses = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Empresa", "Setor", "CNPJ", "Receita", "Margem Líquida",
        "Crescimento", "Dívida", "Caixa", "Dependência Fundador",
        "Valor Equity", "Score Risco", "Maturidade", "Percentil",
        "Plano", "Status", "Criado em",
    ])
    for a in analyses:
        writer.writerow([
            a.company_name, a.sector, a.cnpj or "",
            float(a.revenue), a.net_margin, a.growth_rate or "",
            float(a.debt), float(a.cash), a.founder_dependency,
            float(a.equity_value) if a.equity_value else "",
            a.risk_score or "", a.maturity_index or "", a.percentile or "",
            a.plan.value if a.plan else "", a.status.value,
            a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=analises-valuora.csv"},
    )


# ─── Excel Export (single analysis) ──────────────────────

@router.get("/{analysis_id}/export/xlsx")
async def export_analysis_xlsx(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta os dados de uma análise específica em XLSX (3 abas)."""
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    wb = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1A6B45")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    SUB_FILL   = PatternFill("solid", fgColor="E8F5EF")

    def _style_header_row(ws, row_num, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row_num].height = 22

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    vr = analysis.valuation_result or {}

    # ── Sheet 1: Resumo ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    headers1 = ["Campo", "Valor"]
    ws1.append(headers1)
    _style_header_row(ws1, 1, 2)

    summary_rows = [
        ("Company",          analysis.company_name),
        ("Sector",            analysis.sector),
        ("Tax ID",             analysis.cnpj or "N/A"),
        ("Revenue ($)",     float(analysis.revenue) if analysis.revenue else 0),
        ("Net Margin",   f"{analysis.net_margin * 100:.1f}%"),
        ("Growth",      f"{(analysis.growth_rate or 0) * 100:.1f}%"),
        ("Debt ($)",      float(analysis.debt or 0)),
        ("Cash ($)",       float(analysis.cash or 0)),
        ("Years in business",  analysis.years_in_business or "N/A"),
        ("Employees",        analysis.num_employees or "N/A"),
        ("Equity Value ($)", float(analysis.equity_value) if analysis.equity_value else "N/A"),
        ("Risk Score",   analysis.risk_score or "N/A"),
        ("Maturity",       analysis.maturity_index or "N/A"),
        ("Percentile",        f"{analysis.percentile:.1f}%" if analysis.percentile else "N/A"),
        ("Plano",            analysis.plan.value if analysis.plan else "N/A"),
        ("Status",           analysis.status.value),
        ("Criado em",        analysis.created_at.strftime("%d/%m/%Y %H:%M") if analysis.created_at else ""),
    ]
    for i, row in enumerate(summary_rows, start=2):
        ws1.append(list(row))
        if i % 2 == 0:
            for c in range(1, 3):
                ws1.cell(row=i, column=c).fill = SUB_FILL
    _autofit(ws1)

    # ── Sheet 2: FCF Projetado ────────────────────────────
    ws2 = wb.create_sheet("FCF Projetado")
    fcf_proj = vr.get("fcf_projections", [])
    if fcf_proj:
        fcf_cols = list(fcf_proj[0].keys())
        ws2.append(fcf_cols)
        _style_header_row(ws2, 1, len(fcf_cols))
        for row in fcf_proj:
            ws2.append([row.get(k) for k in fcf_cols])
        _autofit(ws2)
    else:
        ws2["A1"] = "Sem dados de projeção FCF disponíveis."

    # ── Sheet 3: DRE Projetada ────────────────────────────
    ws3 = wb.create_sheet("DRE Projetada")
    pnl_proj = vr.get("pnl_projections", [])
    if pnl_proj:
        pnl_cols = list(pnl_proj[0].keys())
        ws3.append(pnl_cols)
        _style_header_row(ws3, 1, len(pnl_cols))
        for row in pnl_proj:
            ws3.append([row.get(k) for k in pnl_cols])
        _autofit(ws3)
    else:
        ws3["A1"] = "Sem dados de DRE projetada disponíveis."

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = (analysis.company_name or "empresa").replace(" ", "-").lower()
    filename = f"valuation-{safe_name}-{str(analysis.id)[:8]}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Sectors Endpoint ────────────────────────────────────

@router.get("/sectors/list")
async def list_sectors():
    """Retorna todos os setores IBGE disponíveis agrupados. Cached 1h."""
    _cache_key = "qv:static:sectors_list"
    cached = await cache_get(_cache_key)
    if cached:
        return cached
    sectors = get_sector_list()
    groups: Dict = {}
    for s in sectors:
        group = s["group"]
        if group not in groups:
            groups[group] = []
        groups[group].append({"id": s["id"], "label": s["label"]})
    result = {"sectors": sectors, "groups": groups, "total": len(sectors)}
    await cache_set(_cache_key, result, ttl=CACHE_TTL_SHORT)
    return result


# ─── FEEDBACK ────────────────────────────────────────────────────────────────

class FeedbackCreate(BaseModel):
    analysis_id: uuid.UUID
    score: int = Field(..., ge=0, le=10, description="NPS score 0-10")
    comment: Optional[str] = Field(None, max_length=2000)


@router.post("/feedback", status_code=201)
async def submit_feedback(
    payload: FeedbackCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Store NPS/feedback score submitted from the post-analysis modal."""
    feedback = UserFeedback(
        user_id=current_user.id,
        analysis_id=payload.analysis_id,
        score=payload.score,
        comment=payload.comment,
    )
    db.add(feedback)
    await db.commit()
    logger.info(
        f"[FEEDBACK] saved user={current_user.id} analysis={payload.analysis_id} "
        f"score={payload.score} comment={payload.comment!r}"
    )
    return {"message": "Feedback recebido.", "score": payload.score}


# ─── SINGLE ANALYSIS ─────────────────────────────────────────────────────────

@router.get("/{analysis_id}", response_model=AnalysisResponse)
async def get_analysis(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Admins can view any analysis; regular users only their own
    query = select(Analysis).where(
        Analysis.id == analysis_id,
        Analysis.deleted_at.is_(None),
    )
    if not current_user.is_admin:
        query = query.where(Analysis.user_id == current_user.id)
    result = await db.execute(query)
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    return analysis


# ─── Duplicate Analysis ─────────────────────────────────────
@router.post("/{analysis_id}/duplicate", response_model=AnalysisResponse)
async def duplicate_analysis(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a copy of an existing analysis."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
    )
    original = result.scalar_one_or_none()
    if not original:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    
    # Create duplicate with same data but new ID
    duplicate = Analysis(
        user_id=current_user.id,
        company_name=f"{original.company_name} (cópia)",
        sector=original.sector,
        cnpj=original.cnpj,
        revenue=original.revenue,
        net_margin=original.net_margin,
        growth_rate=original.growth_rate,
        debt=original.debt,
        cash=original.cash,
        founder_dependency=original.founder_dependency,
        projection_years=original.projection_years,
        ebitda=original.ebitda,
        recurring_revenue_pct=original.recurring_revenue_pct,
        num_employees=original.num_employees,
        years_in_business=original.years_in_business,
        previous_investment=original.previous_investment,
        qualitative_answers=original.qualitative_answers,
        dcf_weight=original.dcf_weight,
        custom_exit_multiple=original.custom_exit_multiple,
        logo_path=original.logo_path,
        pending_assets=original.pending_assets,
        status=AnalysisStatus.DRAFT,
        valuation_result=None,
        plan=None,
    )
    db.add(duplicate)
    await db.commit()
    await db.refresh(duplicate)
    return duplicate

# ─── Patch Analysis (Archive/Unarchive / Quick Edit) ─────────────────────
@router.patch("/{analysis_id}")
async def patch_analysis(
    analysis_id: uuid.UUID,
    deleted_at: Optional[str] = Body(None, embed=True),
    company_name: Optional[str] = Body(None, embed=True),
    revenue: Optional[float] = Body(None, embed=True),
    net_margin: Optional[float] = Body(None, embed=True),
    ebitda: Optional[float] = Body(None, embed=True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update analysis fields: archive/unarchive or quick-edit basic inputs."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    # ── Quick-edit fields ──────────────────────────────────────
    if company_name is not None:
        stripped = company_name.strip()
        if stripped:
            analysis.company_name = stripped
    if revenue is not None:
        if revenue < 0:
            raise HTTPException(status_code=400, detail="Receita não pode ser negativa.")
        analysis.revenue = revenue
    if net_margin is not None:
        # frontend sends as percentage (e.g. 25.5 → stored as 0.255)
        if net_margin < -100 or net_margin > 100:
            raise HTTPException(status_code=400, detail="Margem líquida deve estar entre -100% e 100%.")
        analysis.net_margin = net_margin / 100.0
    if ebitda is not None:
        analysis.ebitda = ebitda

    # ── Archive / unarchive ────────────────────────────────────
    if deleted_at is not None:
        if deleted_at == "":
            analysis.deleted_at = None
        else:
            try:
                analysis.deleted_at = datetime.now(timezone.utc)
            except Exception:
                raise HTTPException(status_code=400, detail="Formato de data inv\u00e1lido.")
    elif all(v is None for v in (company_name, revenue, net_margin, ebitda)):
        # legacy behaviour: bare PATCH with no fields → unarchive
        analysis.deleted_at = None

    await db.commit()
    # ── Audit trail ────────────────────────────────────────
    changed = [k for k, v in {"company_name": company_name, "revenue": revenue, "net_margin": net_margin, "ebitda": ebitda, "deleted_at": deleted_at}.items() if v is not None]
    await audit_log(
        action="analysis.edit",
        user_id=str(current_user.id),
        user_email=current_user.email,
        resource_id=str(analysis_id),
        detail=f"fields: {','.join(changed)}" if changed else "unarchive",
    )
    return {"message": "Análise atualizada com sucesso."}

# duplicate DELETE /{analysis_id} removed — use /{analysis_id}/permanent for hard delete


# ─── Update Notes ──────────────────────────────────────────
@router.patch("/{analysis_id}/notes")
async def update_notes(
    analysis_id: uuid.UUID,
    notes: Optional[str] = Body(None, embed=True, max_length=50_000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Save user notes for an analysis."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    analysis.notes = notes
    await db.commit()
    await audit_log(
        action="analysis.notes",
        user_id=str(current_user.id),
        user_email=current_user.email,
        resource_id=str(analysis_id),
        detail=f"notes_length={len(notes) if notes else 0}",
    )
    return {"message": "Notas salvas.", "notes": notes}


# ─── Generate Share Token ──────────────────────────────────
import secrets as _secrets
from passlib.context import CryptContext as _CryptContext

_share_pwd_ctx = _CryptContext(schemes=["bcrypt"], deprecated="auto")


class ShareInput(BaseModel):
    password: Optional[str] = None  # if provided, protects the share link
    sections: Optional[List[str]] = None  # subset of allowed sections; None = all


@router.post("/{analysis_id}/share")
async def generate_share_token(
    analysis_id: uuid.UUID,
    body: ShareInput = ShareInput(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate or return existing share token for a public read-only link.
    Optionally password-protect it by providing ``password`` in the request body.
    Calling again with a new password will update it.
    """
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    if not analysis.share_token:
        analysis.share_token = _secrets.token_urlsafe(32)
    if body.password:
        analysis.share_password_hash = _share_pwd_ctx.hash(body.password)
    else:
        # Explicitly passing no password clears any existing protection
        if body.password is not None:
            analysis.share_password_hash = None
    if body.sections is not None:
        # Sanitize: keep only known section keys
        ALLOWED = {"overview", "equity", "projections", "sensitivity", "monte_carlo",
                   "waterfall", "qualitative", "pending_assets", "peers", "ai_analysis"}
        analysis.share_sections = [s for s in body.sections if s in ALLOWED] or None
    await db.commit()
    return {
        "share_token": analysis.share_token,
        "password_protected": bool(analysis.share_password_hash),
        "sections": analysis.share_sections,
    }


# ─── Revoke Share Token ────────────────────────────────────
@router.delete("/{analysis_id}/share")
async def revoke_share_token(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Revoke the share token, making the link inaccessible."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    analysis.share_token = None
    analysis.share_password_hash = None
    await db.commit()
    return {"message": "Link compartilhável revogado."}


# ─── Reanalysis Alert Threshold ────────────────────────────
class AlertInput(BaseModel):
    threshold_pct: Optional[float] = None  # e.g. 0.10 = 10%; None clears the alert


@router.put("/{analysis_id}/alert")
async def set_alert_threshold(
    analysis_id: uuid.UUID,
    body: AlertInput,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Set or clear the reanalysis alert threshold.
    When set, the user will be notified if equity_value changes by >= threshold_pct
    after a re-analysis.
    """
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    if body.threshold_pct is not None and not (0.01 <= body.threshold_pct <= 1.0):
        raise HTTPException(status_code=422, detail="threshold_pct deve estar entre 0.01 e 1.0")
    analysis.reanalysis_alert_pct = body.threshold_pct
    await db.commit()
    return {
        "message": "Alerta configurado." if body.threshold_pct else "Alerta removido.",
        "reanalysis_alert_pct": body.threshold_pct,
    }


# ─── Public Read-Only by Share Token ──────────────────────
@router.get("/public/{share_token}")
async def get_public_analysis(
    share_token: str,
    password: Optional[str] = None,  # DEPRECATED query param — kept for backward compat
    x_share_password: Optional[str] = Header(None),  # preferred: password travels in a header, not the URL
    db: AsyncSession = Depends(get_db),
):
    """Return a read-only summary of an analysis by share token (no auth required).
    If the analysis is password-protected, supply the password via the
    ``X-Share-Password`` header (preferred — never logged in URLs)."""
    pwd = x_share_password or password
    result = await db.execute(
        select(Analysis).where(
            Analysis.share_token == share_token,
            Analysis.deleted_at.is_(None),
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Link inválido ou expirado.")
    if analysis.status != AnalysisStatus.COMPLETED:
        raise HTTPException(status_code=403, detail="Esta análise ainda não foi concluída.")
    # Password check
    if analysis.share_password_hash:
        if not pwd:
            raise HTTPException(status_code=401, detail="Esta análise está protegida por senha.")
        if not _share_pwd_ctx.verify(pwd, analysis.share_password_hash):
            raise HTTPException(status_code=401, detail="Senha incorreta.")

    sections = analysis.share_sections  # None = all
    val_result = analysis.valuation_result or {}
    if sections:
        # Filter valuation_result based on allowed sections
        SECTION_KEYS = {
            "projections": ["fcf_projections", "pnl_projections"],
            "sensitivity": ["sensitivity_table"],
            "monte_carlo": ["monte_carlo"],
            "waterfall": ["waterfall"],
            "qualitative": ["qualitative"],
            "pending_assets": ["pending_assets_adjustment", "equity_value_before_pending"],
            "peers": ["peers", "multiples_valuation"],
        }
        # Build set of keys to KEEP
        keep = {"equity_value_gordon", "equity_value_exit_multiple", "valuation_range",
                "dcf_weight", "multiples_weight"}  # always-on
        for sec in sections:
            for k in SECTION_KEYS.get(sec, []):
                keep.add(k)
        val_result = {k: v for k, v in val_result.items() if k in keep}

    show_ai = (sections is None) or ("ai_analysis" in sections)
    return {
        "company_name": analysis.company_name,
        "sector": analysis.sector,
        "equity_value": float(analysis.equity_value) if analysis.equity_value else None,
        "risk_score": analysis.risk_score,
        "maturity_index": analysis.maturity_index,
        "percentile": analysis.percentile,
        "valuation_result": val_result,
        "ai_analysis": analysis.ai_analysis if show_ai else None,
        "plan": analysis.plan.value if analysis.plan else None,
        "created_at": analysis.created_at.isoformat() if analysis.created_at else None,
        "share_sections": sections,
    }


# ─── Re-análise sem novo pagamento ───────────────────────────────────────────
class ReanalyzeInput(BaseModel):
    """All fields optional — only provided values override the analysis."""
    revenue: Optional[float] = None
    net_margin: Optional[float] = None
    growth_rate: Optional[float] = None
    debt: Optional[float] = None
    cash: Optional[float] = None
    founder_dependency: Optional[float] = None
    projection_years: Optional[int] = None
    ebitda: Optional[float] = None
    recurring_revenue_pct: Optional[float] = None
    num_employees: Optional[int] = None
    years_in_business: Optional[int] = None
    previous_investment: Optional[float] = None
    qualitative_answers: Optional[dict] = None
    dcf_weight: Optional[float] = None
    custom_exit_multiple: Optional[float] = None
    pending_assets: Optional[List[Dict[str, Any]]] = None

    class Config:
        extra = "ignore"


@router.post("/{analysis_id}/reanalyze", response_model=AnalysisResponse)
async def reanalyze(
    analysis_id: uuid.UUID,
    body: ReanalyzeInput,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Re-run the valuation engine with (optionally) updated inputs.

    Requires the analysis to be COMPLETED and already paid (plan set).
    Creates a historical snapshot in AnalysisVersion before overwriting results.
    """
    await _check_user_analysis_rate_limit(str(current_user.id))
    result_row = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
    )
    analysis = result_row.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    if not analysis.plan:
        raise HTTPException(status_code=403, detail="Esta análise ainda não foi paga. Conclua o pagamento primeiro.")
    if analysis.status not in (AnalysisStatus.COMPLETED, AnalysisStatus.FAILED):
        raise HTTPException(status_code=400, detail="Análise não está em estado que permita re-execução.")

    # Snapshot current results before overwriting
    next_version = await db.scalar(
        select(func.coalesce(func.max(AnalysisVersion.version_number), 0) + 1)
        .where(AnalysisVersion.analysis_id == analysis_id)
    )
    if analysis.valuation_result:
        snapshot = AnalysisVersion(
            analysis_id=analysis.id,
            version_number=next_version or 2,
            valuation_result=analysis.valuation_result,
            equity_value=analysis.equity_value,
        )
        db.add(snapshot)

    # Merge provided fields into analysis
    def _v(new_val, old_val):
        return new_val if new_val is not None else old_val

    revenue           = _v(body.revenue,           float(analysis.revenue or 0))
    net_margin        = _v(body.net_margin,         float(analysis.net_margin or 0.10))
    growth_rate       = _v(body.growth_rate,        float(analysis.growth_rate or 0.10))
    debt              = _v(body.debt,               float(analysis.debt or 0))
    cash              = _v(body.cash,               float(analysis.cash or 0))
    founder_dep       = _v(body.founder_dependency, float(analysis.founder_dependency or 0))
    projection_years  = _v(body.projection_years,   analysis.projection_years or 10)
    ebitda            = _v(body.ebitda,             float(analysis.ebitda) if analysis.ebitda else None)
    recurring_rev_pct = _v(body.recurring_revenue_pct, float(analysis.recurring_revenue_pct or 0))
    num_employees     = _v(body.num_employees,      analysis.num_employees or 0)
    years_in_business = _v(body.years_in_business,  analysis.years_in_business or 3)
    prev_investment   = _v(body.previous_investment, float(analysis.previous_investment or 0))
    qual_answers      = _v(body.qualitative_answers, analysis.qualitative_answers)
    dcf_weight        = _v(body.dcf_weight,         float(analysis.dcf_weight) if analysis.dcf_weight else 0.50)
    custom_exit_mult  = _v(body.custom_exit_multiple, analysis.custom_exit_multiple)
    pending_assets_in = _v(body.pending_assets,     analysis.pending_assets)

    analysis.status = AnalysisStatus.PROCESSING

    # Persist updated input fields
    analysis.revenue          = revenue
    analysis.net_margin       = net_margin
    analysis.growth_rate      = growth_rate
    analysis.debt             = debt
    analysis.cash             = cash
    analysis.founder_dependency    = founder_dep
    analysis.projection_years      = projection_years
    analysis.ebitda                = ebitda
    analysis.recurring_revenue_pct = recurring_rev_pct
    analysis.num_employees         = num_employees
    analysis.years_in_business     = years_in_business
    analysis.previous_investment   = prev_investment
    analysis.qualitative_answers   = qual_answers
    analysis.dcf_weight            = dcf_weight
    analysis.custom_exit_multiple  = custom_exit_mult
    analysis.pending_assets        = pending_assets_in
    await db.flush()

    # Try IBGE sector adjustment
    ibge_adj = None
    try:
        cnae_code = _sector_to_cnae(analysis.sector)
        adjustment = await get_dcf_sector_adjustment(
            cnae_code=cnae_code,
            company_revenue=revenue,
            company_growth=growth_rate,
        )
        ibge_adj = adjustment.model_dump()
    except Exception as exc:
        logger.warning("[REANALYZE] IBGE failed: %s — trying DeepSeek fallback", exc)
        try:
            cnae_code = _sector_to_cnae(analysis.sector)
            ai_sector = await estimate_sector_data_with_ai(analysis.sector, cnae_code)
            if ai_sector:
                ibge_adj = ai_sector
        except Exception as exc2:
            logger.warning("[REANALYZE] DeepSeek sector fallback failed: %s", exc2)

    _engine_kwargs = dict(
        years_in_business=years_in_business,
        ebitda=ebitda,
        recurring_revenue_pct=recurring_rev_pct,
        num_employees=num_employees,
        previous_investment=prev_investment,
        qualitative_answers=qual_answers,
        dcf_weight=dcf_weight,
        custom_exit_multiple=custom_exit_mult,
        # v5 diagnostic fields
        company_type=analysis.company_type,
        revenue_ntm=float(analysis.revenue_ntm) if analysis.revenue_ntm else None,
        ebitda_margin=analysis.ebitda_margin,
        tangible_assets=float(analysis.tangible_assets) if analysis.tangible_assets else 0,
        intangible_assets=float(analysis.intangible_assets) if analysis.intangible_assets else 0,
        equity_participations=float(analysis.equity_participations) if analysis.equity_participations else 0,
        monthly_burn_rate=float(analysis.monthly_burn_rate) if analysis.monthly_burn_rate else None,
        pending_assets=pending_assets_in or None,
    )

    try:
        if ibge_adj:
            new_result = await asyncio.to_thread(
                run_valuation_with_ibge,
                revenue=revenue, net_margin=net_margin, sector=analysis.sector,
                ibge_adjustment=ibge_adj, growth_rate=growth_rate,
                debt=debt, cash=cash, founder_dependency=founder_dep,
                projection_years=projection_years, **_engine_kwargs,
            )
        else:
            new_result = await asyncio.to_thread(
                run_valuation,
                revenue=revenue, net_margin=net_margin, sector=analysis.sector,
                growth_rate=growth_rate, debt=debt, cash=cash,
                founder_dependency=founder_dep, projection_years=projection_years,
                **_engine_kwargs,
            )
    except Exception as engine_err:
        logger.error("[REANALYZE] Engine error for %s: %s", analysis_id, engine_err)
        analysis.status = AnalysisStatus.FAILED
        await db.commit()
        raise HTTPException(status_code=500, detail="Erro no motor de valuation durante re-análise. Tente novamente.")

    # Check reanalysis alert threshold
    old_equity = float(analysis.equity_value) if analysis.equity_value else None

    analysis.valuation_result = new_result
    analysis.equity_value  = new_result["equity_value"]
    analysis.risk_score    = new_result["risk_score"]
    analysis.maturity_index = new_result["maturity_index"]
    analysis.percentile    = new_result["percentile"]
    analysis.status        = AnalysisStatus.COMPLETED

    # Fire alert if threshold is configured and value changed significantly
    new_equity = float(analysis.equity_value) if analysis.equity_value else None
    alert_sent = False
    if (
        analysis.reanalysis_alert_pct
        and old_equity
        and new_equity
        and abs(new_equity - old_equity) / old_equity >= analysis.reanalysis_alert_pct
    ):
        _fire_and_forget(send_report_ready_email(
            current_user.email,
            current_user.full_name or current_user.email,
            analysis.company_name,
            f"{settings.FRONTEND_URL}/analise/{analysis.id}",
        ))
        alert_sent = True

    # Invalidate report so a new PDF can be generated on next download
    old_report = (await db.execute(
        select(Report).where(Report.analysis_id == analysis_id)
    )).scalar_one_or_none()
    if old_report:
        await db.delete(old_report)

    # Invalidate cache
    await cache_delete_pattern(f"analysis:{analysis_id}:*")

    await db.commit()
    await db.refresh(analysis)
    # Audit trail for reanalysis
    await audit_log(
        action="analysis.reanalyze",
        user_id=str(current_user.id),
        user_email=current_user.email,
        resource_id=str(analysis_id),
        detail=f"version={next_version}, equity_old={old_equity}, equity_new={new_equity}",
    )
    # Notify user by email (re-run) — only if alert email was NOT already sent
    if not alert_sent:
        _fire_and_forget(send_report_ready_email(
            current_user.email,
            current_user.full_name or current_user.email,
            analysis.company_name,
            f"{settings.FRONTEND_URL}/analise/{analysis.id}",
        ))
    return analysis


# ─── SSE: generation progress ────────────────────────────────────────────────
@router.get("/{analysis_id}/generation-progress")
async def generation_progress_stream(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """SSE stream: polls Redis for PDF generation progress every 2 s.

    The key ``gen_progress:{analysis_id}`` is set by
    ``_generate_and_send_report`` in payments.py.
    """
    from sse_starlette.sse import EventSourceResponse
    import asyncio, json

    # Verify ownership
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    key = f"gen_progress:{analysis_id}"

    async def event_gen():
        import time
        max_ticks = 150  # 150 × 2 s = 5 min
        start_time = time.monotonic()
        for _ in range(max_ticks):
            data = await cache_get(key)
            if data:
                pct = data.get("pct", 0)
                if pct and pct > 0:
                    elapsed = time.monotonic() - start_time
                    estimated_total = elapsed / (pct / 100)
                    data["eta_seconds"] = max(0, round(estimated_total - elapsed))
                yield {"event": "progress", "data": json.dumps(data)}
                if data.get("done") or data.get("error"):
                    return
            else:
                yield {"event": "waiting", "data": "{}"}
            await asyncio.sleep(2)
        yield {"event": "timeout", "data": "{}"}

    return EventSourceResponse(event_gen())


# ─── REST: generation status (poll-friendly alternative) ─────────────────────
@router.get("/{analysis_id}/generation-status")
async def generation_status(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return current generation progress from Redis (poll every 2s from frontend)."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    key = f"gen_progress:{analysis_id}"
    data = await cache_get(key)
    if data:
        return data
    return {"step": 0, "message": "Aguardando início…", "pct": 0, "done": False, "error": None}


# ─── PDF download ─────────────────────────────────────────────────────────────
@router.get("/{analysis_id}/pdf")
async def download_analysis_pdf(
    analysis_id: uuid.UUID,
    theme: str = "light",
    white_label: bool = False,
    lang: str = "pt",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Serve the generated PDF report directly as a file download.

    Query params:
        theme: 'light' (default) | 'dark' — visual style of the report
        white_label: when true and the user is associated to a partner, the PDF
            is co-branded with partner's company name and brand color.
        lang: 'pt' (default) | 'en' — report language
    """
    if theme not in ("light", "dark"):
        theme = "light"
    if lang not in ("pt", "en"):
        lang = "pt"
    # Allow analysis owner OR admin to download
    query = select(Analysis).where(Analysis.id == analysis_id)
    if not current_user.is_admin:
        query = query.where(Analysis.user_id == current_user.id)

    result = await db.execute(query)
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    # Resolve partner branding when requested
    partner_branding = None
    if white_label and getattr(current_user, "partner_id", None):
        from app.models.models import Partner as _Partner
        pres = await db.execute(select(_Partner).where(_Partner.id == current_user.partner_id))
        partner = pres.scalar_one_or_none()
        if partner:
            partner_branding = {
                "company_name": partner.company_name or "",
                "brand_color": partner.brand_color or "#10b981",
                "logo_path": None,
            }

    # When variants requested, always regenerate (cached file would be default theme)
    force_regen = (theme == "dark") or (partner_branding is not None) or (lang == "en")

    report_result = await db.execute(
        select(Report).where(Report.analysis_id == analysis_id)
    )
    report = report_result.scalar_one_or_none()

    if not report:
        # No report record: allowed if analysis is paid+completed (e.g. after reanalyze
        # deletes the old report so a fresh PDF can be generated on next download).
        if not analysis.plan or analysis.status != AnalysisStatus.COMPLETED:
            raise HTTPException(
                status_code=404,
                detail="Relatório PDF ainda não gerado. Aguarde a confirmação do pagamento.",
            )
        # Generate PDF on-the-fly and persist a new Report record
        if not analysis.valuation_result:
            raise HTTPException(status_code=404, detail="Dados de valuation não encontrados para gerar o PDF.")
        from app.services.pdf_service import generate_report_pdf
        from app.core.security import create_download_token
        # Pre-generate UUID so verify badge URL can be embedded in the PDF before persisting
        report_uuid = uuid.uuid4()
        try:
            pdf_path = await asyncio.to_thread(
                generate_report_pdf, analysis,
                False, "", None, partner_branding, theme, lang,
                str(report_uuid),  # verify_id — embeds badge URL in footer
            )
        except Exception as exc:
            logging.getLogger(__name__).error("PDF gen failed for %s: %s", analysis_id, exc)
            raise HTTPException(status_code=500, detail="Falha ao gerar o PDF. Tente novamente.")
        report = Report(
            id=report_uuid,
            analysis_id=analysis.id,
            version=1,
            file_path=pdf_path if lang == "pt" else None,
            file_path_en=pdf_path if lang == "en" else None,
            download_token=create_download_token(str(analysis.id)),
        )
        db.add(report)
        await db.commit()
        await db.refresh(report)

    # If PDF file is missing on disk (e.g. Railway ephemeral FS), or a variant is requested, regenerate
    # For EN lang, check file_path_en first
    cached_path = report.file_path_en if lang == "en" else report.file_path
    if force_regen or not cached_path or not os.path.exists(cached_path):
        if not analysis.valuation_result:
            raise HTTPException(
                status_code=404,
                detail="Dados de valuation não encontrados para regenerar o PDF.",
            )
        from app.services.pdf_service import generate_report_pdf
        try:
            pdf_path = await asyncio.to_thread(
                generate_report_pdf, analysis,
                False, "", None, partner_branding, theme, lang,
                str(report.id),  # verify_id
            )
        except Exception as exc:
            logging.getLogger(__name__).error("PDF regen failed for %s: %s", analysis_id, exc)
            raise HTTPException(status_code=500, detail="Falha ao regenerar o PDF.")
        # Only persist file_path for default variant (pt, light, no partner); variants are throwaway
        is_default_variant = (lang == "pt" and theme == "light" and partner_branding is None)
        is_en_default = (lang == "en" and theme == "light" and partner_branding is None)
        if is_default_variant:
            report.file_path = pdf_path
            await db.commit()
        elif is_en_default:
            report.file_path_en = pdf_path
            await db.commit()
        else:
            # Serve variant directly without overriding canonical file_path
            company = (analysis.company_name or str(analysis_id)).replace(" ", "_")
            suffix = f"-{theme}" if theme == "dark" else ""
            if partner_branding:
                suffix += "-cobranded"
            if lang == "en":
                suffix += "-en"
            return FileResponse(
                pdf_path,
                media_type="application/pdf",
                filename=f"relatorio-valuora-{company}{suffix}.pdf",
            )
        cached_path = pdf_path

    company = (analysis.company_name or str(analysis_id)).replace(" ", "_")
    lang_suffix = "-en" if lang == "en" else ""
    return FileResponse(
        cached_path,
        media_type="application/pdf",
        filename=f"relatorio-valuora-{company}{lang_suffix}.pdf",
    )


# ─── Favorites ────────────────────────────────────────────────────────────────

@router.get("/favorites/list")
async def list_favorites(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return all favorited analysis IDs for the current user."""
    rows = (await db.execute(
        select(UserFavorite.analysis_id)
        .where(UserFavorite.user_id == current_user.id)
    )).scalars().all()
    return {"favorites": [str(r) for r in rows]}


@router.post("/{analysis_id}/favorite", response_model=dict)
async def add_favorite(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Toggle-add: favorite an analysis."""
    # Verify ownership
    analysis = (await db.execute(
        select(Analysis).where(Analysis.id == analysis_id, Analysis.user_id == current_user.id)
    )).scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    # Idempotent insert
    existing = (await db.execute(
        select(UserFavorite).where(
            UserFavorite.user_id == current_user.id,
            UserFavorite.analysis_id == analysis_id,
        )
    )).scalar_one_or_none()
    if not existing:
        fav = UserFavorite(user_id=current_user.id, analysis_id=analysis_id)
        db.add(fav)
        await db.commit()
    return {"favorited": True, "analysis_id": str(analysis_id)}


@router.delete("/{analysis_id}/favorite", response_model=dict)
async def remove_favorite(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Remove an analysis from favorites."""
    existing = (await db.execute(
        select(UserFavorite).where(
            UserFavorite.user_id == current_user.id,
            UserFavorite.analysis_id == analysis_id,
        )
    )).scalar_one_or_none()
    if existing:
        await db.delete(existing)
        await db.commit()
    return {"favorited": False, "analysis_id": str(analysis_id)}


# ─── Version History ──────────────────────────────────────────────────────────

@router.get("/{analysis_id}/versions")
async def get_versions(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return all version snapshots for an analysis."""
    # Verify ownership
    analysis = (await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )).scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    versions = (await db.execute(
        select(AnalysisVersion)
        .where(AnalysisVersion.analysis_id == analysis_id)
        .order_by(AnalysisVersion.version_number.desc())
    )).scalars().all()

    def _extract_params(vr: dict) -> dict:
        """Extract key comparable parameters from a valuation result."""
        params = vr.get("parameters", {})
        return {
            "growth_rate": round(float(params.get("growth_rate", 0)) * 100, 2),
            "net_margin": round(float(params.get("net_margin", 0)) * 100, 2),
            "wacc": round(float(vr.get("wacc", 0)) * 100, 2),
            "revenue": float(params.get("revenue", 0)),
            "projection_years": int(params.get("projection_years", 10)),
            "dlom_pct": round(float((vr.get("dlom") or {}).get("dlom_pct", 0)) * 100, 1),
            "risk_score": float(vr.get("risk_score", vr.get("survival", {}).get("risk_score", 0))),
            "tv_percentage": float(vr.get("tv_percentage", 0)),
        }

    current_params = _extract_params(analysis.valuation_result or {})

    return {
        "analysis_id": str(analysis_id),
        "company_name": analysis.company_name,
        "current_value": float(analysis.equity_value) if analysis.equity_value else None,
        "current_params": current_params,
        "versions": [
            {
                "id": str(v.id),
                "version_number": v.version_number,
                "equity_value": float(v.equity_value) if v.equity_value else None,
                "created_at": v.created_at.isoformat() if v.created_at else None,
                "params": _extract_params(v.valuation_result or {}),
            }
            for v in versions
        ],
    }


# ─── Valuation History (Valuation E) ─────────────────────

@router.get("/valuation-history")
async def get_valuation_history(
    company_name: str = Query(..., description="Nome da empresa para buscar histórico"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return valuation history for a company name — shows equity value over time."""
    # Escape LIKE wildcards to prevent injection
    safe_name = company_name.replace("%", "\\%").replace("_", "\\_")
    analyses = (await db.execute(
        select(Analysis)
        .where(
            Analysis.user_id == current_user.id,
            Analysis.company_name.ilike(f"%{safe_name}%"),
            Analysis.status == AnalysisStatus.COMPLETED,
            Analysis.deleted_at.is_(None),
            Analysis.equity_value.is_not(None),
        )
        .order_by(Analysis.created_at.asc())
        .limit(200)
    )).scalars().all()

    return {
        "company_name": company_name,
        "history": [
            {
                "id": str(a.id),
                "equity_value": float(a.equity_value),
                "risk_score": float(a.risk_score) if a.risk_score else None,
                "maturity_index": float(a.maturity_index) if a.maturity_index else None,
                "wacc": a.valuation_result.get("wacc") if a.valuation_result else None,
                "growth_rate": a.valuation_result.get("parameters", {}).get("growth_rate") if a.valuation_result else None,
                "created_at": a.created_at.isoformat() if a.created_at else None,
                "label": a.created_at.strftime("%b/%Y") if a.created_at else "",
            }
            for a in analyses
        ],
        "count": len(analyses),
    }


# ─── M&A Comparables (Valuation H) ───────────────────────

@router.get("/{analysis_id}/ma-comparables")
async def get_analysis_ma_comparables(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get M&A comparable transactions for a company's sector — AI-powered."""
    analysis = (await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )).scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    cache_key = f"qv:ma_comparables:{analysis.sector}"
    cached = await cache_get(cache_key)
    if cached:
        return cached

    revenue = float(analysis.revenue) if analysis.revenue else 0
    data = await get_ma_comparables(sector=analysis.sector or "tecnologia", revenue=revenue)
    if not data:
        raise HTTPException(status_code=503, detail="Não foi possível obter comparáveis de M&A no momento.")

    # Cache for 7 days
    await cache_set(cache_key, data, ttl=604800)
    return data


# ─── Inverse Projection (F6) ──────────────────────────────

@router.post("/inverse-projection")
async def inverse_projection(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Given an analysis and a target equity value, find the revenue growth rate
    and/or net margin needed to reach that target.  Binary-search over the
    driving variable while holding all other parameters constant.
    """
    analysis_id = payload.get("analysis_id")
    target_equity = float(payload.get("target_equity", 0))
    variable = payload.get("variable", "growth_rate")  # growth_rate | net_margin
    lo = float(payload.get("range_min", 0.0))
    hi = float(payload.get("range_max", 1.5))

    if not analysis_id or target_equity <= 0:
        raise HTTPException(status_code=422, detail="analysis_id e target_equity são obrigatórios.")

    analysis = (await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.status == AnalysisStatus.COMPLETED,
        )
    )).scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada ou não concluída.")
    if not analysis.plan:
        raise HTTPException(status_code=403, detail="Projeção inversa disponível apenas para análises com plano ativo.")

    vr = analysis.valuation_result or {}
    params = vr.get("parameters", {})

    async def _value_at(x: float) -> float:
        """Run valuation with variable set to x; return equity value."""
        overrides = dict(params)
        overrides[variable] = x
        _revenue = overrides.get("revenue", float(analysis.revenue or 0))
        _net_margin = overrides.get("net_margin", float(analysis.net_margin or 0))
        _growth_rate = overrides.get("growth_rate", float(analysis.growth_rate or 0))
        _sector = analysis.sector or "servicos"
        _years_in_business = int(overrides.get("years_in_business", analysis.years_in_business or 5))
        _projection_years = int(overrides.get("projection_years", 10))
        _qualitative_answers = analysis.qualitative_answers or {}
        _debt = float(overrides.get("debt", 0))
        result = await asyncio.to_thread(
            lambda: run_valuation_with_ibge(
                revenue=_revenue,
                net_margin=_net_margin,
                growth_rate=_growth_rate,
                sector=_sector,
                years_in_business=_years_in_business,
                projection_years=_projection_years,
                qualitative_answers=_qualitative_answers,
                debt=_debt,
            )
        )
        return float(result.get("equity_value_final") or result.get("equity_value", 0))

    # Binary search — max 20 iterations, ~0.1% precision
    results = []
    for _ in range(20):
        mid = (lo + hi) / 2
        val = await _value_at(mid)
        results.append({"x": round(mid * 100, 2), "equity": round(val, 0)})
        if abs(val - target_equity) / max(target_equity, 1) < 0.001:
            break
        if val < target_equity:
            lo = mid
        else:
            hi = mid

    solution_x = results[-1]["x"] if results else None
    solution_val = results[-1]["equity"] if results else None

    # Build a small curve for charting: 10 evenly spaced points — fetched in parallel
    curve = []
    range_lo = float(payload.get("range_min", 0.0))
    range_hi = float(payload.get("range_max", 1.5))
    step = (range_hi - range_lo) / 9
    xs = [range_lo + step * i for i in range(10)]
    curve_values = await asyncio.gather(*[_value_at(x) for x in xs])
    for x, v in zip(xs, curve_values):
        curve.append({"x": round(x * 100, 1), "equity": round(v, 0)})

    label = "Taxa de Crescimento" if variable == "growth_rate" else "Margem Líquida"
    return {
        "variable": variable,
        "variable_label": label,
        "target_equity": target_equity,
        "solution_x_pct": solution_x,
        "solution_equity": solution_val,
        "curve": curve,
        "current_x_pct": round(float(params.get(variable, 0)) * 100, 2),
        "current_equity": float(analysis.equity_value or 0),
    }


# ─── Portfolio Excel Export ──────────────────────────────
@router.get("/export-portfolio-xlsx")
async def export_portfolio_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download all paid analyses as a multi-sheet Excel workbook."""
    from io import BytesIO
    from app.services.portfolio_xlsx_service import generate_portfolio_xlsx
    from app.models.models import SimulationLog

    # Fetch all paid analyses for the user
    res = await db.execute(
        select(Analysis)
        .where(Analysis.user_id == current_user.id, Analysis.deleted_at == None, Analysis.plan != None)  # noqa: E711
        .order_by(Analysis.created_at.desc())
    )
    analyses_db = res.scalars().all()

    # Fetch simulations grouped by analysis_id
    sim_res = await db.execute(
        select(SimulationLog).where(
            SimulationLog.analysis_id.in_([a.id for a in analyses_db])
        ).order_by(SimulationLog.created_at.desc())
    )
    sims_db = sim_res.scalars().all()

    simulations_by_id: dict = {}
    for s in sims_db:
        key = str(s.analysis_id)
        if key not in simulations_by_id:
            simulations_by_id[key] = []
        simulations_by_id[key].append({
            "equity_value": float(s.equity_value) if s.equity_value else None,
            "parameters": s.parameters or {},
            "created_at": s.created_at.isoformat() if s.created_at else "",
        })

    analyses_data = [
        {
            "id": str(a.id),
            "company_name": a.company_name,
            "sector": a.sector,
            "plan": a.plan.value if hasattr(a.plan, "value") else str(a.plan),
            "equity_value": float(a.equity_value) if a.equity_value else None,
            "risk_score": float(a.risk_score) if a.risk_score else None,
            "maturity_index": float(a.maturity_index) if a.maturity_index else None,
            "revenue": float(a.revenue) if a.revenue else None,
            "net_margin": float(a.net_margin) if a.net_margin else None,
            "growth_rate": float(a.growth_rate) if a.growth_rate else None,
            "status": a.status.value if hasattr(a.status, "value") else str(a.status),
            "created_at": a.created_at.isoformat() if a.created_at else "",
            "valuation_result": a.valuation_result or {},
        }
        for a in analyses_db
    ]

    xlsx_bytes = generate_portfolio_xlsx(analyses_data, simulations_by_id)
    from datetime import date
    filename = f"valuora-portfolio-{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Confirm generate — user explicitly triggers PDF generation ───────────────
@router.post("/{analysis_id}/confirm-generate")
async def confirm_generate(
    analysis_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    User clicks "Gerar relatório definitivamente".
    Marks generate_confirmed=True, sets generated_at and triggers PDF generation.
    Can only be called once — raises 400 if already confirmed.
    Requires a PAID payment for this analysis.
    """
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    if analysis.generate_confirmed:
        raise HTTPException(status_code=400, detail="Este relatório já foi gerado e não pode ser alterado.")

    if not analysis.plan:
        raise HTTPException(
            status_code=400,
            detail="Pagamento ainda não confirmado. Aguarde a confirmação do pagamento antes de gerar.",
        )

    if analysis.status != AnalysisStatus.COMPLETED:
        raise HTTPException(
            status_code=400,
            detail="A análise ainda está sendo processada. Aguarde a conclusão.",
        )

    # Mark as confirmed
    analysis.generate_confirmed = True
    analysis.generated_at = datetime.now(timezone.utc)
    await db.commit()

    # Trigger PDF generation in background (same as payment webhook)
    from app.routes.payments import _generate_and_send_report
    background_tasks.add_task(_generate_and_send_report, str(analysis.id), str(current_user.id))

    return {"message": "Geração iniciada. Você receberá o PDF por e-mail em aproximadamente 90 segundos."}


# ─── Report error — user reports a problem with the generated report ──────────
class ReportErrorRequest(BaseModel):
    description: str = Field(..., min_length=5, max_length=1000)


@router.post("/{analysis_id}/report-error")
async def report_error(
    analysis_id: uuid.UUID,
    body: ReportErrorRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    User reports a problem with their generated report.
    Sends an email to the admin and returns the WhatsApp link.
    """
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    if not analysis.generate_confirmed:
        raise HTTPException(status_code=400, detail="Relatório ainda não foi gerado.")

    # Send email to admin
    from app.services.email_service import send_email
    admin_email = settings.ADMIN_EMAIL or "giovannesartor@gmail.com"
    plan_label = analysis.plan.value.capitalize() if analysis.plan else "N/A"

    html_body = f"""
    <h2>🔴 Reporte de erro em relatório de valuation</h2>
    <table style="border-collapse:collapse;width:100%">
      <tr><td style="padding:8px;font-weight:bold">ID da Análise</td><td style="padding:8px">{analysis_id}</td></tr>
      <tr style="background:#f9f9f9"><td style="padding:8px;font-weight:bold">Empresa</td><td style="padding:8px">{analysis.company_name}</td></tr>
      <tr><td style="padding:8px;font-weight:bold">Plano</td><td style="padding:8px">{plan_label}</td></tr>
      <tr style="background:#f9f9f9"><td style="padding:8px;font-weight:bold">Cliente</td><td style="padding:8px">{current_user.full_name or "—"} ({current_user.email})</td></tr>
      <tr><td style="padding:8px;font-weight:bold">Gerado em</td><td style="padding:8px">{analysis.generated_at.strftime('%d/%m/%Y %H:%M') if analysis.generated_at else '—'}</td></tr>
      <tr style="background:#f9f9f9"><td style="padding:8px;font-weight:bold">Descrição do erro</td><td style="padding:8px">{body.description}</td></tr>
    </table>
    <br>
    <p><a href="https://valuora.online/admin/analises/{analysis_id}" style="background:#059669;color:white;padding:10px 20px;text-decoration:none;border-radius:8px">View and fix analysis in admin →</a></p>
    """

    try:
        await send_email(
            to_email=admin_email,
            subject=f"🔴 Erro no relatório: {analysis.company_name} (plano {plan_label})",
            html_body=html_body,
        )
    except Exception as e:
        logger.warning("[REPORT-ERROR] Email failed: %s", e)

    # Build WhatsApp link with pre-filled message
    wa_message = (
        f"Olá, encontrei um problema no meu relatório do Valuora.\n"
        f"Empresa: {analysis.company_name}\n"
        f"ID: {analysis_id}\n"
        f"Problema: {body.description}"
    )
    import urllib.parse
    wa_link = f"https://wa.me/5554999536435?text={urllib.parse.quote(wa_message)}"

    return {
        "message": "Erro reportado. Nossa equipe foi notificada e entrará em contato.",
        "whatsapp_url": wa_link,
    }


# ─── VL IA — Chat with your valuation ────────────────────────────────────────

class ChatMessage(BaseModel):
    message: str = Field(..., min_length=1, max_length=1500)
    history: list[dict] = Field(default_factory=list)

CHAT_SYSTEM_PROMPT = """You are Valuora AI, a valuation assistant for Valuora.
Seu papel é ajudar o dono/sócio da empresa a entender o laudo de valuation da empresa deles.
Seja direto, objetivo e use linguagem acessível (não jargão excessivo).
When quoting values, use USD ($) formatting.
Não invente dados — baseie todas respostas nos dados reais abaixo.
Não faça valuation novo — use apenas os resultados já calculados.
Limite suas respostas a no máximo 4 parágrafos curtos.

DADOS DA EMPRESA E VALUATION:
{context}
"""

@router.post("/{analysis_id}/chat")
async def chat_with_analysis(
    analysis_id: uuid.UUID,
    body: ChatMessage,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """VL IA — chat assistant for paid valuations only. Uses DeepSeek with analysis context."""
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")
    if not analysis.plan:
        raise HTTPException(status_code=403, detail="O chat de IA está disponível apenas para análises com plano ativo.")

    from app.services.deepseek_service import call_deepseek
    import json as _json

    vr = analysis.valuation_result or {}
    mc = vr.get("monte_carlo", {})
    peers = vr.get("peers", {})

    context = (
        f"Empresa: {analysis.company_name}\n"
        f"Setor: {analysis.sector}\n"
        f"Annual revenue: $ {float(analysis.revenue or 0):,.2f}\n"
        f"Margem líquida: {(float(analysis.net_margin or 0) * 100):.1f}%\n"
        f"Taxa de crescimento: {(float(analysis.growth_rate or 0) * 100):.1f}%\n"
        f"Final Equity Value: $ {float(analysis.equity_value or 0):,.2f}\n"
        f"Enterprise Value: $ {float(vr.get('enterprise_value_gordon', 0) or 0):,.2f}\n"
        f"Score de Risco: {float(analysis.risk_score or 0):.0f}/100\n"
        f"Índice de Maturidade: {float(analysis.maturity_index or 0):.0f}/100\n"
        f"Ke (Custo de Capital): {float(vr.get('ke', 0) or 0) * 100:.1f}%\n"
        f"DLOM (Desconto liquidez): {float((vr.get('dlom') or {}).get('rate', 0) or 0) * 100:.1f}%\n"
f"Monte Carlo P25: $ {float(mc.get('p25', 0) or 0):,.2f}\n"
f"Monte Carlo P50: $ {float(mc.get('p50', 0) or 0):,.2f}\n"
f"Monte Carlo P75: $ {float(mc.get('p75', 0) or 0):,.2f}\n"
        f"Plano: {analysis.plan.value if analysis.plan else 'N/A'}\n"
    )
    if analysis.ai_analysis:
        context += f"\nAnálise estratégica (resumo):\n{analysis.ai_analysis[:800]}\n"

    system = CHAT_SYSTEM_PROMPT.format(context=context)

    # Build messages list: system + history (last 6 turns) + new user message
    history = (body.history or [])[-6:]
    messages = [{"role": "system", "content": system}]
    for h in history:
        role = h.get("role")
        content = h.get("content")
        if role in ("user", "assistant") and isinstance(content, str) and content.strip():
            messages.append({"role": role, "content": content[:800]})
    messages.append({"role": "user", "content": body.message})

    try:
        reply = await call_deepseek("", max_tokens=600, messages_override=messages)
    except Exception as e:
        logger.warning("[CHAT] DeepSeek error: %s", e)
        raise HTTPException(status_code=502, detail="Serviço de IA temporariamente indisponível. Tente novamente.")

    return {"reply": reply}


# ─── Inconsistency Check ──────────────────────────────────────────────────────

@router.post("/{analysis_id}/check-inconsistencies")
async def check_inconsistencies(
    analysis_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Run DeepSeek inconsistency detection on the analysis financial data.
    Returns a list of warnings (may be empty). No auth required beyond ownership.
    """
    result = await db.execute(
        select(Analysis).where(
            Analysis.id == analysis_id,
            Analysis.user_id == current_user.id,
            Analysis.deleted_at.is_(None),
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Análise não encontrada.")

    financial_data = {
        "revenue": float(analysis.revenue or 0),
        "net_margin": float(analysis.net_margin or 0),
        "growth_rate": float(analysis.growth_rate or 0),
        "debt": float(analysis.debt or 0),
        "cash": float(analysis.cash or 0),
        "ebitda": float(analysis.ebitda or 0) if hasattr(analysis, "ebitda") else None,
        "sector": analysis.sector,
        "years_in_business": getattr(analysis, "years_in_business", None),
    }
    if analysis.extracted_data:
        for k in ("net_income", "gross_profit", "operating_expenses", "cogs"):
            if analysis.extracted_data.get(k) is not None:
                financial_data[k] = analysis.extracted_data[k]

    data = await detect_data_inconsistencies(financial_data, analysis.sector or "geral")
    return data
